"""Tests for returns/sku_grain.py and the returns companion's By-SKU tab.

Run: python returns/tests/test_sku_grain.py

Synthetic frames, shaped like prep()'s outputs, so each rule can be isolated:
Retail/Trade never blended, ROW returns omitted rather than zeroed, the rate
floor, the exchange convention, and the movement columns.
"""
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, "..", ".."))

import pandas as pd
from returns import sku_grain

FAILURES = []


def check(label, got, want, tol=1e-9):
    if isinstance(want, float) or isinstance(got, float):
        ok = (got is None and want is None) or (
            got is not None and want is not None and abs(got - want) <= tol)
    else:
        ok = got == want
    print(f"  {'PASS' if ok else 'FAIL'}  {label}: got {got!r}, want {want!r}")
    if not ok:
        FAILURES.append(label)


def frames(row_returns=False):
    """One SKU, four markets/segments.

    UK Retail: 30 orders, 3 returned (10%) -- above the 20-order floor
    UK Trade:   5 orders, 1 returned      -- below the floor
    US Retail: 25 orders, 5 returned (20%)
    ROW Retail: 8 orders, 0 returned      -- sales real, returns unrecordable
    """
    rows = []
    for i in range(30):
        rows.append(("SKU-A", f"UK-R-{i}", "UK", "Retail", 2, 100.0))
    for i in range(5):
        rows.append(("SKU-A", f"UK-T-{i}", "UK", "Trade", 10, 400.0))
    for i in range(25):
        rows.append(("SKU-A", f"US-R-{i}", "US", "Retail", 1, 80.0))
    for i in range(8):
        rows.append(("SKU-A", f"ROW-R-{i}", "ROW", "Retail", 1, 90.0))
    s = pd.DataFrame(rows, columns=["sku", "order", "mkt", "seg", "units", "cash"])

    ret_rows = [("SKU-A", f"UK-R-{i}", "UK", "Retail", 2, False) for i in range(3)]
    ret_rows += [("SKU-A", "UK-T-0", "UK", "Trade", 4, False)]
    ret_rows += [("SKU-A", f"US-R-{i}", "US", "Retail", 1, i == 0) for i in range(5)]
    if row_returns:
        ret_rows += [("SKU-A", "ROW-R-0", "ROW", "Retail", 1, False)]
    ret = pd.DataFrame(ret_rows, columns=["sku", "order", "mkt", "seg", "qty", "is_exchange"])

    # Value frame: refund value per returned order. The US order US-R-0 is an
    # exchange, so its value must be excluded from cash but its units/orders
    # still counted.
    # Sign convention: refund_val is `-sum(Line: Total)` over Refund Line rows,
    # whose Total is already negative -- so value arrives POSITIVE. (build.py's
    # _sku_return_value applies .abs() defensively; the companion's
    # _sku_aggregate doesn't, and relies on this.)
    v_rows = [("SKU-A", f"UK-R-{i}", "UK", "Retail", 2, 200.0, False) for i in range(3)]
    v_rows += [("SKU-A", "UK-T-0", "UK", "Trade", 4, 1600.0, False)]
    v_rows += [("SKU-A", f"US-R-{i}", "US", "Retail", 1, 80.0, i == 0) for i in range(5)]
    shopv = pd.DataFrame(v_rows, columns=["sku", "order", "mkt", "seg", "qty", "val", "is_exch_line"])
    return s, ret, shopv


def test_retail_and_trade_are_separate():
    print("test_retail_and_trade_are_separate")
    s, ret, shopv = frames()
    cuts, meta = sku_grain.aggregate(s, ret, shopv)
    a = cuts["SKU-A"]
    check("retail orders", a["retail"]["orders"], 30 + 25 + 8)
    check("trade orders", a["trade"]["orders"], 5)
    check("retail returned orders", a["retail"]["returned_orders"], 8)
    check("trade returned orders", a["trade"]["returned_orders"], 1)
    check("blended orders = retail + trade", a["blended"]["orders"],
          a["retail"]["orders"] + a["trade"]["orders"])
    # The lock: the headline cut is retail, and it must NOT contain trade.
    check("retail excludes trade returns",
          a["retail"]["units_returned"], 3 * 2 + 5 * 1)


def test_row_returns_omitted_not_zeroed():
    print("test_row_returns_omitted_not_zeroed")
    s, ret, shopv = frames(row_returns=False)
    cuts, meta = sku_grain.aggregate(s, ret, shopv)
    row = cuts["SKU-A"]["row"]
    check("row recordable flag", meta["row_returns_recordable"], False)
    check("row sales orders present (real)", row["orders"], 8)
    check("row return_rate absent, not 0.0", "return_rate" in row, False)
    check("row returned_orders absent, not 0", "returned_orders" in row, False)
    ok = "omitted, not zero" in meta["row_note"]
    print(f"  {'PASS' if ok else 'FAIL'}  meta note explains the omission")
    if not ok:
        FAILURES.append("row note")


def test_row_returns_appear_when_source_has_them():
    """The omission must be data-driven: give the source one ROW return and
    the columns come back without a code change.
    """
    print("test_row_returns_appear_when_source_has_them")
    s, ret, shopv = frames(row_returns=True)
    cuts, meta = sku_grain.aggregate(s, ret, shopv)
    row = cuts["SKU-A"]["row"]
    check("row recordable flag", meta["row_returns_recordable"], True)
    check("row returned_orders now present", row["returned_orders"], 1)


def test_rate_floor():
    print("test_rate_floor")
    s, ret, shopv = frames()
    cuts, meta = sku_grain.aggregate(s, ret, shopv, floor=20)
    a = cuts["SKU-A"]
    check("uk_retail rate (30 orders, above floor)", a["uk_retail"]["return_rate"], 3 / 30)
    check("uk_trade rate withheld (5 orders)", a["uk_trade"]["return_rate"], None)
    check("uk_trade counts still present", a["uk_trade"]["returned_orders"], 1)
    check("uk_trade flagged below floor", a["uk_trade"]["below_floor"], True)
    lowered = sku_grain.aggregate(s, ret, shopv, floor=1)[0]["SKU-A"]
    check("floor is a parameter, not baked in", lowered["uk_trade"]["return_rate"], 1 / 5)


def test_exchange_convention():
    """Units and orders include exchanges; cash excludes exchange value."""
    print("test_exchange_convention")
    s, ret, shopv = frames()
    cuts, _ = sku_grain.aggregate(s, ret, shopv)
    us = cuts["SKU-A"]["us_retail"]
    check("us returned orders include the exchange", us["returned_orders"], 5)
    check("us units returned include the exchange", us["units_returned"], 5)
    # 5 refund lines at 80 = 400 gross; one is an exchange, so 320 net.
    check("us cash excludes exchange value", us["value_returned"], 320.0)
    check("value is positive (refund_val sign convention)", us["value_returned"] > 0, True)


def test_uplift():
    print("test_uplift")
    cur = {"value_returned": 150.0, "units_returned": 10, "return_rate": 0.2}
    prior = {"value_returned": 100.0, "units_returned": 20, "return_rate": None}
    check("value uplift +50%", sku_grain.uplift(cur, prior, "value_returned"), 0.5)
    check("units uplift -50%", sku_grain.uplift(cur, prior, "units_returned"), -0.5)
    check("suppressed prior rate -> None", sku_grain.uplift(cur, prior, "return_rate"), None)
    check("no prior at all -> None", sku_grain.uplift(cur, None, "value_returned"), None)
    check("zero prior base -> None",
          sku_grain.uplift(cur, {"value_returned": 0.0}, "value_returned"), None)


def test_tab_writes():
    print("test_tab_writes")
    from openpyxl import Workbook, load_workbook
    from returns.excel_companion import _build_by_sku
    s, ret, shopv = frames()
    cuts, meta = sku_grain.aggregate(s, ret, shopv)
    prior = {"SKU-A": {"retail": {"value_returned": 400.0, "units_returned": 8,
                                    "return_rate": 0.12}}}
    attrs = {"SKU-A": {"description": "Test Handle", "category": "Handle",
                        "subcategory": "T Bar", "department": "Cabinetry",
                        "finish": "Brass", "family": "TESTER", "status": "Live"}}
    wb = Workbook(); wb.remove(wb.active)
    ws = _build_by_sku(wb, None, "Jul 2026", cuts=cuts, meta=meta, prior_cuts=prior,
                        attrs_by_sku=attrs)
    path = "/tmp/returns_by_sku.xlsx"
    wb.save(path)
    ws = load_workbook(path)["By SKU"]
    headers = [ws.cell(4, i).value for i in range(1, ws.max_column + 1)]
    bands = [ws.cell(3, i).value for i in range(1, ws.max_column + 1) if ws.cell(3, i).value]
    print(f"  {ws.max_column} columns, {ws.max_row} rows")
    print(f"  bands: {bands}")
    ok = any("transparency only" in str(b) for b in bands)
    print(f"  {'PASS' if ok else 'FAIL'}  blended block labelled transparency-only")
    if not ok:
        FAILURES.append("blended label")
    ok = any("sales only" in str(b) for b in bands)
    print(f"  {'PASS' if ok else 'FAIL'}  ROW block labelled sales-only")
    if not ok:
        FAILURES.append("row label")
    # 10 cuts, 9 of which carry a rate -- ROW's is absent because its returns
    # side was omitted. Checked positionally too, so a future reordering can't
    # satisfy the count while putting a rate back inside the ROW band.
    check("rate columns = cuts with a returns side",
          sum(1 for h in headers if h == "Return Rate"), 9)
    row_band_cols = [i for i in range(1, ws.max_column + 1)
                     if ws.cell(3, i).value and "ROW" in str(ws.cell(3, i).value)]
    row_start = row_band_cols[0]
    row_headers = []
    for i in range(row_start, ws.max_column + 1):
        if i > row_start and ws.cell(3, i).value:
            break
        row_headers.append(headers[i - 1])
    print(f"  ROW band headers: {row_headers}")
    ok = "Return Rate" not in row_headers and "Returns Cash \u00a3" not in row_headers
    print(f"  {'PASS' if ok else 'FAIL'}  ROW band carries no returns columns")
    if not ok:
        FAILURES.append("row band columns")
    check("description written", ws.cell(5, 3).value, "Test Handle")
    note = str(ws.cell(ws.max_row, 1).value)
    ok = "20 orders" in note and "Retail is the headline" in note
    print(f"  {'PASS' if ok else 'FAIL'}  note states the floor and the headline basis")
    if not ok:
        FAILURES.append("note")
    print(f"  wrote {path}")


def test_narrow_fallback_still_works():
    """An un-updated caller passing only sku_agg must still get a sheet."""
    print("test_narrow_fallback_still_works")
    from openpyxl import Workbook
    from returns.excel_companion import _build_by_sku
    sku_agg = pd.DataFrame([{"sku": "SKU-A", "department": "Cabinetry", "family": "TESTER",
                              "finish": "Brass", "value_returned": 100.0,
                              "units_returned": 5, "return_rate": 0.1}])
    wb = Workbook(); wb.remove(wb.active)
    ws = _build_by_sku(wb, sku_agg, "Jul 2026")
    check("narrow width", ws.max_column, 8)


if __name__ == "__main__":
    for fn in (test_retail_and_trade_are_separate, test_row_returns_omitted_not_zeroed,
               test_row_returns_appear_when_source_has_them, test_rate_floor,
               test_exchange_convention, test_uplift, test_tab_writes,
               test_narrow_fallback_still_works):
        fn()
        print()
    if FAILURES:
        print(f"FAILED: {len(FAILURES)} check(s): {FAILURES}")
        sys.exit(1)
    print("all checks passed")
