"""Build the returns dashboard's Excel companion -- values-only, matching the
visual format of Plank_Q1_2026_Returns_Companion_REAL.xlsx (measured off that
file; same design system as trading/excel_companion.py's, reused from
common/excel_styling.py).

Unlike trading, there's no pre-built "contract" artifact for returns to read --
this module calls build.prep() itself and aggregates the same way
render.build_cube() does, just collapsed to sku grain (no month/market/seg
breakout) since the companion reports the whole period at once.

Real improvement over the reference file, not just a reformat: the reference
file's ROW figures are all "PENDING" -- Q1 2026's source (the old workbook) had
no ship-to country field at all. Since 2026-08-13, returns/build.py derives mkt
from Shipping: Country Code the same way trading does, so ROW is real data here,
not a placeholder. The Reconciliation & Data Quality tab says so explicitly
rather than silently dropping the caveat.
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pandas as pd
from openpyxl import Workbook
from returns import build
from openpyxl.utils import get_column_letter
from common.excel_styling import (
    title_row, subtitle_row, section_header, table_header, data_row, total_row,
    kpi_block, note_row, block_label, FMT_CURRENCY, FMT_PERCENT, FMT_PERCENT_SIGNED,
    FMT_INT,
)
from returns import sku_grain


def _sku_aggregate(s, ret, shopv):
    """Per-SKU totals for the whole period (no month/market/seg breakout) --
    same fields render.build_cube() computes, collapsed to sku grain, plus
    finish/department/family carried from s's own per-sku taxonomy (assumed
    1:1 with sku, same assumption build_cube() makes)."""
    sales_agg = s.groupby("sku").agg(units_sold=("units", "sum"), gross_sales=("cash", "sum"),
                                       orders=("order", "nunique"))
    ret_agg = ret.groupby("sku").agg(units_returned=("qty", "sum"), returned_orders=("order", "nunique"))
    d = shopv[shopv.qty != 0]
    exch = d[d.is_exch_line].groupby("sku")["val"].sum()
    stock = d.groupby("sku")["val"].sum()
    value_returned = stock - exch.reindex(stock.index).fillna(0)
    taxonomy = s.drop_duplicates("sku").set_index("sku")[["finish", "department", "family", "status"]]

    agg = (sales_agg.join(ret_agg, how="left")
                     .join(value_returned.rename("value_returned"), how="left")
                     .join(taxonomy))
    agg = agg.fillna({"units_returned": 0, "returned_orders": 0, "value_returned": 0})
    agg["return_rate"] = (agg["returned_orders"] / agg["orders"]).where(agg["orders"] > 0, 0)
    return agg.reset_index()


def _build_overview(wb, headline, by_market, period_label, source_note):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16

    title_row(ws, 1, f"{period_label} - Returns Report (Companion)", 8)
    subtitle_row(ws, 2, "REAL DATA - single-count  |  orders-based rate  |  ex-VAT", 8)
    note_row(ws, 3, source_note, 8)

    section_header(ws, 5, f"{period_label} Headline", 8)
    kpi_block(ws, 6, "Return Rate (orders-based)", f"{headline['order_rate']:.1%}",
               f"{headline['returned_orders']:,.0f} of {headline['orders']:,.0f} orders", 1)
    kpi_block(ws, 6, "Returns Value", f"£{headline['value_returned']:,.0f}", "ex-VAT", 3)
    kpi_block(ws, 6, "Return Rate (units)", f"{headline['unit_rate']:.1%}",
               f"{headline['units_returned']:,.0f} of {headline['units_sold']:,.0f} units", 5)
    kpi_block(ws, 6, "Returned Orders", f"{headline['returned_orders']:,.0f}", "single-count", 7)

    uk, us, row = by_market.loc["UK"], by_market.loc["US"], by_market.loc["ROW"]
    kpi_block(ws, 10, "\U0001F1EC\U0001F1E7 UK Orders Returned", f"{uk['return_rate']:.1%}",
               f"{uk['returned_orders']:,.0f} of {uk['orders']:,.0f}", 1)
    kpi_block(ws, 10, "\U0001F1FA\U0001F1F8 US Orders Returned", f"{us['return_rate']:.1%}",
               f"{us['returned_orders']:,.0f} of {us['orders']:,.0f}", 3)
    kpi_block(ws, 10, "\U0001F30D ROW Orders Returned",
               f"{row['return_rate']:.1%}" if row["orders"] else "n/a",
               f"{row['returned_orders']:,.0f} of {row['orders']:,.0f}" if row["orders"] else "no ROW orders this period", 5)
    kpi_block(ws, 10, "Units Returned", f"{headline['units_returned']:,.0f}", "single-count", 7)

    section_header(ws, 15, "Returns by Country - ORDERS  (reconciliation key)", 4)
    table_header(ws, 16, ["Country", "Orders", "Orders returned", "% returned"])
    r = 17
    names = {"UK": "United Kingdom", "US": "United States", "ROW": "Rest of World (ROW)"}
    for k in ["UK", "US", "ROW"]:
        row_data = by_market.loc[k]
        data_row(ws, r, [names[k], row_data["orders"], row_data["returned_orders"], row_data["return_rate"]],
                  [None, FMT_INT, FMT_INT, FMT_PERCENT])
        r += 1
    tot = by_market.loc["Total"]
    total_row(ws, r, ["TOTAL", tot["orders"], tot["returned_orders"], tot["return_rate"]],
               [None, FMT_INT, FMT_INT, FMT_PERCENT])
    r += 1
    note_row(ws, r, "Country here is ship-to (Shipping: Country Code), the same derivation trading "
                     "uses for ROW -- real data, not a placeholder (see Reconciliation tab for the "
                     "one caveat: ROW's own overlap isn't asserted, since ReturnZap's Country field "
                     "is GB/US only).", 4)


def _style_drill_level(ws, row, level, ncols=6):
    from common.excel_styling import DRILL_FILLS, DRILL_INDENT, DRILL_BOLD
    from openpyxl.styles import Alignment, PatternFill, Font
    fill = PatternFill("solid", fgColor=DRILL_FILLS[level])
    font_color = "FFFFFF" if level == 0 else "1F2933"
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Arial", size=10, bold=DRILL_BOLD[level], color=font_color)
        c.fill = fill
        if col == 1:
            c.alignment = Alignment(horizontal="left", vertical="center", indent=DRILL_INDENT[level])
        else:
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].outline_level = level
    if level == 3:
        ws.row_dimensions[row].hidden = True


def _build_drill_down(wb, sku_agg, headline, period_label):
    ws = wb.create_sheet("Returns Drill-Down")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    for col, w in zip("ABCDEF", [34, 20, 15, 14, 14, 12]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - Returns Drill-Down", 6)
    subtitle_row(ws, 2, "Total -> Product Type -> Collection -> SKU (returns cash)  |  click +/- "
                         "to expand  |  ex-VAT", 6)
    note_row(ws, 3, "SKUs collapsed by default - use the outline +/- to drill in. Children sum "
                     "to parents.", 6)

    headers = ["Product / Collection / SKU", "Type / Finish", "Returns cash",
               "Units returned", "Returned orders", "Return rate"]
    table_header(ws, 5, headers)
    fmts = [None, None, FMT_CURRENCY, FMT_INT, FMT_INT, FMT_PERCENT]

    r = 6
    data_row(ws, r, ["TOTAL", "", headline["value_returned"], headline["units_returned"],
                      headline["returned_orders"], headline["order_rate"]], fmts)
    _style_drill_level(ws, r, 0)
    r += 1

    by_dept = sku_agg.groupby("department").agg(
        value_returned=("value_returned", "sum"), units_returned=("units_returned", "sum"),
        returned_orders=("returned_orders", "sum"), orders=("orders", "sum"), n_families=("family", "nunique"))
    for dept, drow in by_dept.sort_values("value_returned", ascending=False).iterrows():
        drate = drow["returned_orders"] / drow["orders"] if drow["orders"] else 0
        data_row(ws, r, [dept or "(uncategorised)", f"{int(drow['n_families'])} collections",
                          drow["value_returned"], drow["units_returned"], drow["returned_orders"], drate], fmts)
        _style_drill_level(ws, r, 1)
        r += 1

        fam_group = sku_agg[sku_agg["department"] == dept].groupby("family").agg(
            value_returned=("value_returned", "sum"), units_returned=("units_returned", "sum"),
            returned_orders=("returned_orders", "sum"), orders=("orders", "sum"))
        for fam, frow in fam_group.sort_values("value_returned", ascending=False).iterrows():
            frate = frow["returned_orders"] / frow["orders"] if frow["orders"] else 0
            data_row(ws, r, [fam or "(uncategorised)", dept, frow["value_returned"], frow["units_returned"],
                              frow["returned_orders"], frate], fmts)
            _style_drill_level(ws, r, 2)
            r += 1

            skus = sku_agg[(sku_agg["department"] == dept) & (sku_agg["family"] == fam)]
            for _, srow in skus.sort_values("value_returned", ascending=False).iterrows():
                data_row(ws, r, [srow["sku"], srow["finish"], srow["value_returned"], srow["units_returned"],
                                  srow["returned_orders"], srow["return_rate"]], fmts)
                _style_drill_level(ws, r, 3)
                r += 1

    ws.freeze_panes = "A6"
    ws.sheet_format.outlineLevelRow = 3


def _build_detail(wb, reason_mix_df, by_status, by_finish, period_label):
    ws = wb.create_sheet("Detail")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [26, 16, 14, 12, 12]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - Returns Detail", 5)
    subtitle_row(ws, 2, "by reason / by status / by finish  |  ex-VAT", 5)

    r = 4
    section_header(ws, r, "By Return Reason (unit counts)", 3); r += 1
    table_header(ws, r, ["Reason", "Returned units", "Share"]); r += 1
    total_units = reason_mix_df["units_returned"].sum()
    for reason, row_data in reason_mix_df.iterrows():
        data_row(ws, r, [reason, row_data["units_returned"], row_data["share"]],
                  [None, FMT_INT, FMT_PERCENT])
        r += 1
    total_row(ws, r, ["TOTAL", total_units, 1.0], [None, FMT_INT, FMT_PERCENT]); r += 2

    section_header(ws, r, "By Product Status (returns cash & orders)", 5); r += 1
    table_header(ws, r, ["Status", "Returns cash", "Units returned", "Returned orders", "% orders"]); r += 1
    fmts = [None, FMT_CURRENCY, FMT_INT, FMT_INT, FMT_PERCENT]
    for status, row_data in by_status.iterrows():
        if status == "Total":
            continue
        data_row(ws, r, [status, row_data.get("value_returned", 0), row_data["units_returned"],
                          row_data["returned_orders"], row_data["return_rate"]], fmts)
        r += 1
    tot = by_status.loc["Total"]
    total_row(ws, r, ["TOTAL", tot.get("value_returned", 0), tot["units_returned"],
                       tot["returned_orders"], tot["return_rate"]], fmts)
    r += 2

    section_header(ws, r, "By Finish (returns cash & rate)", 5); r += 1
    table_header(ws, r, ["Finish", "Returns cash", "Units returned", "Returned orders", "% orders"]); r += 1
    for finish, row_data in by_finish.iterrows():
        if finish == "Total":
            continue
        data_row(ws, r, [finish, row_data.get("value_returned", 0), row_data["units_returned"],
                          row_data["returned_orders"], row_data["return_rate"]], fmts)
        r += 1


# ── By SKU: segment x market grain ───────────────────────────────────────
# Built 2026-08-13 to give the returns companion the same depth the trading
# companion's By-SKU tab got, adapted for three things that make returns
# different (all three explained at length in returns/sku_grain.py):
#
#   * Retail leads and Trade sits beside it, un-blended (build.py §5.3 is a
#     LOCKED decision). The combined figure is the last block and is labelled
#     as transparency-only, matching how run() treats by_month_blended.
#   * The ROW block's returns columns are OMITTED when the source can't record
#     ROW returns -- not written as zero, which would read as "ROW never
#     returns anything". ROW sales columns are real and always shown.
#   * Return rate is withheld below the 20-order floor; the order counts that
#     make up the rate are always shown, so nothing is hidden, only unrounded
#     noise is.
#
# Ranking is by ALL returns cash (Retail + Trade), not Retail alone: the
# question this tab answers is "what is coming back", and a Trade-heavy SKU
# ranked to the bottom of a 700-row sheet is a miss. The lock is about not
# blending a headline metric, not about sort order -- and every headline
# metric on the row is still segment-separated.

# (metric_key, header, format) for each cut block. One list, so header count,
# format count and value count cannot drift apart.
_CUT_METRICS = [
    ("value_returned",  "Returns Cash \u00a3", FMT_CURRENCY),
    ("units_returned",  "Units Returned",       FMT_INT),
    ("returned_orders", "Returned Orders",      FMT_INT),
    ("orders",          "Orders",               FMT_INT),
    ("return_rate",     "Return Rate",          FMT_PERCENT),
    ("value_rate",      "Returns % of Sales",   FMT_PERCENT),
]
# The ROW block keeps only the sales-side facts when returns aren't recordable.
_CUT_METRICS_SALES_ONLY = [
    ("orders",     "Orders",     FMT_INT),
    ("units_sold", "Units Sold", FMT_INT),
]
_ATTRS = [
    ("rank",        "Rank",                None),
    ("sku",         "SKU",                 None),
    ("description", "Product Description", None),
    ("category",    "Product Category",    None),
    ("department",  "Product Type",        None),
    ("subcategory", "Sub Category",        None),
    ("finish",      "Finish",              None),
    ("family",      "Family/Collection",   None),
    ("status",      "Status",              None),
]


def _by_sku_layout(meta, has_prior):
    """[(block_label, [(cut_key_or_None, metric_key, header, fmt), ...]), ...]

    Computed rather than hardcoded because the column set genuinely varies:
    the ROW block narrows when ROW returns aren't recordable, and the movement
    block disappears when no prior period was loaded. Everything downstream
    (headers, formats, block bands, values) is derived from this one structure.
    """
    blocks = [("", [(None, k, h, f) for k, h, f in _ATTRS])]
    for key in sku_grain.CUT_KEYS:
        label = sku_grain.CUT_LABELS[key]
        if key in sku_grain.ROW_DEPENDENT_CUTS and not meta["row_returns_recordable"]:
            metrics = _CUT_METRICS_SALES_ONLY
            label = f"{label} (sales only)"
        else:
            metrics = _CUT_METRICS
        if key == "blended":
            label = f"{label} - transparency only, not the headline"
        blocks.append((label, [(key, mk, h, f) for mk, h, f in metrics]))
    if has_prior:
        blocks.append(("vs PRIOR PERIOD (Retail)", [
            ("retail", "_vs_value", "Returns Cash vs prior", FMT_PERCENT_SIGNED),
            ("retail", "_vs_units", "Units Returned vs prior", FMT_PERCENT_SIGNED),
            ("retail", "_prior_value", "Prior Returns Cash \u00a3", FMT_CURRENCY),
            ("retail", "_prior_rate", "Prior Return Rate", FMT_PERCENT),
        ]))
    return blocks


def _build_by_sku(wb, sku_agg, period_label, cuts=None, meta=None, prior_cuts=None,
                   attrs_by_sku=None):
    """The By-SKU tab. Falls back to the original 8-column view when no cut
    data is passed, so an existing caller that hasn't been updated still
    produces a working sheet rather than an error.
    """
    ws = wb.create_sheet("By SKU")
    ws.sheet_view.showGridLines = False

    if not cuts:
        for col, w in zip("ABCDEFGH", [8, 20, 15, 18, 15, 14, 14, 12]):
            ws.column_dimensions[col].width = w
        title_row(ws, 1, f"{period_label} - Returns by SKU (full)", 8)
        subtitle_row(ws, 2, "ranked by returns cash  |  ex-VAT  |  filterable", 8)
        table_header(ws, 3, ["Rank", "SKU", "Type", "Collection", "Finish",
                              "Returns cash", "Units returned", "Return rate"])
        fmts = [FMT_INT, None, None, None, None, FMT_CURRENCY, FMT_INT, FMT_PERCENT]
        ranked = sku_agg.sort_values("value_returned", ascending=False)
        r = 4
        for i, (_, row_data) in enumerate(ranked.iterrows(), start=1):
            data_row(ws, r, [i, row_data["sku"], row_data["department"], row_data["family"],
                              row_data["finish"], row_data["value_returned"],
                              row_data["units_returned"], row_data["return_rate"]], fmts)
            r += 1
        ws.freeze_panes = "A4"
        return ws

    blocks = _by_sku_layout(meta, bool(prior_cuts))
    flat = [(blk, spec) for blk, specs in blocks for spec in specs]
    headers = [spec[2] for _, spec in flat]
    fmts = [spec[3] for _, spec in flat]
    ncols = len(headers)

    for i, w in enumerate([8, 22, 34, 16, 15, 16, 14, 16, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(_ATTRS) + 1, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    title_row(ws, 1, f"{period_label} - Returns by SKU", ncols)
    subtitle_row(ws, 2, "ranked by all returns cash (Retail + Trade)  |  ex-VAT  |  "
                        "order-month cohort  |  Retail and Trade never blended into a headline  |  "
                        "filterable", ncols)
    col = 1
    for label, specs in blocks:
        if label:
            block_label(ws, 3, label, col, col + len(specs) - 1)
        col += len(specs)
    table_header(ws, 4, headers)

    def rank_key(sku):
        b = cuts.get(sku, {}).get("blended", {})
        return -(b.get("value_returned") or 0)

    r = 5
    for i, sku in enumerate(sorted(cuts, key=rank_key), start=1):
        by_cut = cuts[sku]
        attrs = (attrs_by_sku or {}).get(sku, {})
        prior_retail = (prior_cuts or {}).get(sku, {}).get("retail")
        values = []
        for _, (cut_key, metric_key, _, _) in flat:
            if cut_key is None:
                values.append(i if metric_key == "rank" else
                              (sku if metric_key == "sku" else attrs.get(metric_key)))
                continue
            m = by_cut.get(cut_key)
            if metric_key == "_vs_value":
                values.append(sku_grain.uplift(m, prior_retail, "value_returned"))
            elif metric_key == "_vs_units":
                values.append(sku_grain.uplift(m, prior_retail, "units_returned"))
            elif metric_key == "_prior_value":
                values.append((prior_retail or {}).get("value_returned"))
            elif metric_key == "_prior_rate":
                values.append((prior_retail or {}).get("return_rate"))
            else:
                values.append((m or {}).get(metric_key))
        data_row(ws, r, values, fmts)
        r += 1

    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{r - 1}"

    note = (f"{len(cuts)} SKUs. Return rate is returned orders / orders and is withheld where a "
            f"cut has fewer than {meta['floor']} orders -- the order counts are shown either side "
            f"of it, so the rate is the only thing suppressed. Order and unit figures include "
            f"exchanges; cash figures exclude exchange-attributable value. Retail is the headline "
            f"basis; the Retail + Trade block is shown for transparency and is not a headline "
            f"figure. {meta['row_note']}")
    if prior_cuts is None:
        note += (" Movement columns are omitted: no prior period was loaded to compare against.")
    note_row(ws, r + 1, note, ncols)
    return ws


def _build_reconciliation(wb, headline, by_market, period_label, headline_blended=None):
    ws = wb.create_sheet("Reconciliation & Data Quality")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50

    title_row(ws, 1, f"{period_label} - Returns Reconciliation & Data Quality", 3)
    subtitle_row(ws, 2, "what reconciles, what does not, and why", 3)
    note_row(ws, 3, "This tab is deliberately loud about gaps - the reconciliation contract requires it.", 3)

    r = 5
    section_header(ws, r, "ROW bucket", 3); r += 1
    row_data = by_market.loc["ROW"]
    data_row(ws, r, ["ROW orders", int(row_data["orders"]),
                      "real data (ship-to country) -- resolved 2026-08-13, was PENDING in the old source"])
    r += 1
    data_row(ws, r, ["ROW returned orders", int(row_data["returned_orders"]),
                      "overlap NOT asserted for ROW: ReturnZap's own Country field is GB/US only, "
                      "so a ROW return can never be confirmed against it -- see build.py's prep()"])
    r += 1
    note_row(ws, r, "RESOLVED vs. the old source: country reconciliation now uses real ship-to data "
                     "for sales; the one remaining caveat is the overlap check, not the country split "
                     "itself.", 3)
    r += 2

    section_header(ws, r, "Headline consistency", 3); r += 1
    table_header(ws, r, ["Source", "Returns cash", "Note"]); r += 1
    data_row(ws, r, [f"Pipeline headline (this companion, {headline.get('basis', 'Retail')})",
                      headline["value_returned"],
                      "single source (ReturnZap sheet, deduped) -- no cross-view drift possible, "
                      "unlike the old multi-workbook source"], [None, FMT_CURRENCY, None])
    r += 1
    if headline_blended:
        # Both figures on the page, on purpose. Until 2026-08-13 this companion
        # headlined the blended figure while the HTML dashboard headlined
        # Retail, so the two deliverables disagreed for the same period. The
        # companion now follows the dashboard and §5.3; showing what the
        # blended view says makes the size of that correction visible instead
        # of leaving readers of an earlier file to wonder why the number moved.
        data_row(ws, r, ["Retail + Trade blended (NOT the headline)",
                          headline_blended["value_returned"],
                          "shown for transparency only. Locked decision (build.py §5.3): the "
                          "headline is Retail; Retail and Trade are never combined into one rate "
                          "or value. Companions issued before 2026-08-13 headlined this blended "
                          "figure in error -- corrected, not restated silently."],
                  [None, FMT_CURRENCY, None])
        r += 1
        data_row(ws, r, ["Headline order rate (Retail)", headline["order_rate"],
                          f"{int(headline['returned_orders'])} returned orders of "
                          f"{int(headline['orders'])} Retail orders"], [None, FMT_PERCENT, None])
        r += 1
        data_row(ws, r, ["Blended order rate (NOT the headline)", headline_blended["order_rate"],
                          f"{int(headline_blended['returned_orders'])} of "
                          f"{int(headline_blended['orders'])} orders, Retail and Trade combined"],
                  [None, FMT_PERCENT, None])
        r += 1
    r += 1

    section_header(ws, r, "Notes", 3); r += 1
    notes = [
        "Source: the same ReturnZap Drive sheet + rolling Matrixify snapshot the HTML dashboard "
        "itself uses (returns/build.py's prep()) -- not a separate re-derivation.",
        "Single-count and orders-based headline rate, per the locked returns methodology.",
        "Headline basis is RETAIL, matching the HTML dashboard and build.py \u00a75.3. Trade is "
        "computed and reported, never blended into the headline.",
        "Order-month cohort throughout: a return is counted in the month the ORDER was placed, "
        "not the month it came back.",
        "Returns reported separately, never netted into revenue.",
        "Values-only (no formulas).",
    ]
    for n in notes:
        note_row(ws, r, f"\u2022  {n}", 3)
        r += 1


def build_returns_companion(out_path, period_label, sales_df, ld_std, returns_df, month_nums, year,
                             source_note="", prior_month_nums=None, prior_year=None):
    """Build the companion.

    prior_month_nums / prior_year (2026-08-13): the window to compare the
    By-SKU tab's movement columns against. Optional -- there is no committed
    returns contract to chain to, so a comparative means re-running prep() on
    the prior window, which only works if that period's orders are present in
    `sales_df`. Omitted or absent -> the movement columns are left out and the
    tab says so, rather than showing a -100% drop off a period that was never
    loaded.
    """
    s, ret, zap, shopv, months = build.prep(sales_df, ld_std, returns_df, month_nums, year)

    by_market = build.by_group(s, ret, "mkt", ["UK", "US", "ROW"])
    fins = s["finish"].value_counts().index.tolist()
    by_finish = build.by_group(s, ret, "finish", fins)
    from returns.build import STAT
    by_status = build.by_group(s, ret, "status", STAT)
    reason = build.reason_mix(zap)
    vsplit = build.value_split(shopv)

    sku_agg = _sku_aggregate(s, ret, shopv)
    # attach value_returned onto the by_group blocks (by_group itself doesn't carry $ value)
    for block, col in [(by_finish, "finish"), (by_status, "status")]:
        vmap = sku_agg.groupby(col)["value_returned"].sum()
        block["value_returned"] = block.index.map(vmap).fillna(0)

    # ── Headline basis, corrected 2026-08-13 ────────────────────────────────
    # This block used to read by_market.loc["Total"], i.e. Retail AND Trade
    # blended, while run() -- which feeds the HTML dashboard -- headlines
    # by_month(s_retail, ret_retail). Same period, two deliverables, two
    # different headline return rates, and the Excel one was the side
    # contradicting build.py §5.3 ("headline defaults to RETAIL; trade is
    # computed and reported separately; the two are never combined into one
    # blended rate/value"), which is a LOCKED decision.
    #
    # The companion now headlines Retail, matching the dashboard and the lock.
    # Trade and blended are still computed and still reported -- on the
    # Reconciliation tab, labelled -- so nothing is lost, and the difference
    # between this month's figure and previously-issued companions is a
    # correction with a stated cause, not a silent restatement.
    s_retail, ret_retail = s[s.seg == "Retail"], ret[ret.seg == "Retail"]
    retail_orders = s_retail["order"].nunique()
    retail_returned = ret_retail["order"].nunique()
    retail_units_sold = s_retail["units"].sum()
    retail_units_returned = ret_retail["qty"].sum()
    vsplit_retail = build.value_split(shopv, "Retail")

    headline = {
        "orders": retail_orders, "returned_orders": retail_returned,
        "order_rate": (retail_returned / retail_orders) if retail_orders else 0,
        "units_sold": retail_units_sold, "units_returned": retail_units_returned,
        "unit_rate": (retail_units_returned / retail_units_sold) if retail_units_sold else 0,
        "value_returned": vsplit_retail["stock_value"],
        "basis": "Retail",
    }
    # Kept for the Reconciliation tab: what the blended view would say, so the
    # two are visibly different numbers rather than one quietly replacing the
    # other.
    tot = by_market.loc["Total"]
    headline_blended = {
        "orders": tot["orders"], "returned_orders": tot["returned_orders"],
        "order_rate": tot["return_rate"],
        "value_returned": vsplit["stock_value"],
    }

    # ── SKU grain (2026-08-13) ──────────────────────────────────────────────
    cuts, cuts_meta = sku_grain.aggregate(s, ret, shopv)
    prior_cuts, prior_reason = (None, "no prior period requested")
    if prior_month_nums:
        prior_cuts, prior_reason = sku_grain.prior_period_aggregate(
            sales_df, ld_std, returns_df, prior_month_nums, prior_year or year)
    if prior_reason and prior_month_nums:
        print(f"returns companion: {prior_reason}", file=sys.stderr)

    descriptions = {}
    try:
        descriptions = build.load_line_detail_names()
    except Exception as e:
        print(f"returns companion: product descriptions unavailable ({type(e).__name__}: {e}) -- "
              f"the By-SKU tab's description column will be empty", file=sys.stderr)
    ld_idx = ld_std.drop_duplicates("sku").set_index("sku")
    tax = s.drop_duplicates("sku").set_index("sku")
    attrs_by_sku = {}
    for sku in cuts:
        attrs_by_sku[sku] = {
            "description": descriptions.get(sku),
            "category": ld_idx["category"].get(sku) if "category" in ld_idx.columns else None,
            "subcategory": ld_idx["subcategory"].get(sku) if "subcategory" in ld_idx.columns else None,
            "department": tax["department"].get(sku) if "department" in tax.columns else None,
            "finish": tax["finish"].get(sku) if "finish" in tax.columns else None,
            "family": tax["family"].get(sku) if "family" in tax.columns else None,
            "status": tax["status"].get(sku) if "status" in tax.columns else None,
        }

    wb = Workbook()
    _build_overview(wb, headline, by_market, period_label,
                     source_note or f"Source: {period_label} committed returns build. Values-only.")
    _build_drill_down(wb, sku_agg, headline, period_label)
    _build_detail(wb, reason, by_status, by_finish, period_label)
    _build_by_sku(wb, sku_agg, period_label, cuts=cuts, meta=cuts_meta,
                   prior_cuts=prior_cuts, attrs_by_sku=attrs_by_sku)
    _build_reconciliation(wb, headline, by_market, period_label,
                           headline_blended=headline_blended)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
