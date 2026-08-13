"""Shared styling for the trading dashboard's Excel companion.

Every constant here was measured directly off Plank_Q2_2026_Trade_Companion_REAL.xlsx
(the reference file) via openpyxl -- not guessed. Values-only workbook (no formulas),
matching that file's own stated convention.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Colors (measured from the reference file) ---
NAVY = "13293D"          # title bar fill
SECTION_BLUE = "1F4E5F"  # section header fill
HEADER_GREY = "E8EEF1"   # table header / TOTAL row fill
TEXT_DARK = "1F2933"     # body text
TEXT_GREY = "6B7280"     # KPI label / subtext
WHITE = "FFFFFFFF"

# Drill-down depth shading (Trade Drill-Down tab), lightest for deepest level
DRILL_FILLS = {0: SECTION_BLUE, 1: "DCE6EB", 2: "EEF3F5", 3: "F5F7F8"}
DRILL_INDENT = {0: 0, 1: 1, 2: 3, 3: 5}
DRILL_BOLD = {0: True, 1: True, 2: False, 3: False}

# --- Number formats ---
FMT_CURRENCY = r"\£#,##0"
FMT_PERCENT = "0.0%"
FMT_PERCENT_SIGNED = r"\+0.0%;\-0.0%;0.0%"
FMT_INT = "#,##0"

THIN = Side(style="thin")


def title_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26


def subtitle_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=9, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")


def section_header(ws, row, text, ncols, start_col=1):
    if ncols > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=SECTION_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")


def table_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=TEXT_DARK)
        c.fill = PatternFill("solid", fgColor=HEADER_GREY)
        c.alignment = Alignment(
            horizontal="right" if i > 0 else "left", vertical="center", wrap_text=True
        )
        c.border = Border(bottom=THIN)


def data_row(ws, row, values, formats=None, start_col=1, bold=False, fill=None):
    """values: list; formats: optional list of number_format strings aligned to values
    (None entries left as General, used for text columns)."""
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = Font(name="Arial", size=10, bold=bold, color=TEXT_DARK)
        c.alignment = Alignment(horizontal="right" if i > 0 else "left", vertical="center")
        if formats and formats[i]:
            c.number_format = formats[i]
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)


def total_row(ws, row, values, formats=None, start_col=1):
    data_row(ws, row, values, formats, start_col, bold=True, fill=HEADER_GREY)
    for i in range(len(values)):
        ws.cell(row=row, column=start_col + i).border = Border(top=THIN)


def kpi_block(ws, row, label, value_str, subtext, col):
    """A 3-row KPI block (label / big number / subtext) starting at `row`, in
    the given column pair (col, col+1 merged) -- matches the reference file's
    Summary tab layout exactly (e.g. A6:B6 label, A7:B7... value spans just A)."""
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = Font(name="Arial", size=9, bold=True, color=TEXT_GREY)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)

    val = ws.cell(row=row + 1, column=col, value=value_str)
    val.font = Font(name="Arial", size=18, bold=True, color=NAVY)

    sub = ws.cell(row=row + 2, column=col, value=subtext)
    sub.font = Font(name="Arial", size=9, color=TEXT_GREY)
    sub.border = Border(bottom=THIN)
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
    ws.cell(row=row + 2, column=col + 1).border = Border(bottom=THIN)


def note_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=8, italic=True, color=TEXT_GREY)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
