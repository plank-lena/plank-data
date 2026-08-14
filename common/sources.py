"""Single source-of-truth for every external data connection the trading and
returns builders read from -- Drive-hosted sheets, Matrixify order exports,
and the local snapshot paths those connectors land at. One place to update
when a source moves: get the new Drive share link -> copy its file ID out of
the URL (.../d/<FILE_ID>/edit...) -> replace the one line in SOURCES below ->
re-run `python common/sources.py preflight`.

IMPORTANT -- who actually calls the connectors: the Google Drive and
Matrixify MCP tools are only reachable from an interactive Claude Code
session (or the shared Claude Project), never from a plain `python` process
-- there is no local API key this module could use to call them itself, by
design (see CLAUDE.md's "no .env" rule for these sources). So this module
does NOT perform any network fetch. What it owns:
  - the pinned IDs/tabs/expected shape (SOURCES below) -- the thing that
    changes when a source moves;
  - where a fetched snapshot lands on disk (the *_SNAPSHOT paths below) --
    every builder reads from these paths unchanged, whether the snapshot was
    just refreshed live or is a few days stale;
  - normalizing a raw connector download into the exact shape the EXISTING
    parsers already expect, so no parser gains source-branching logic (the
    pattern trading/line_detail.py's own docstring already describes:
    "dropbox... just needs to land a file shaped like the committed local
    snapshot... so no parsing logic ever branches on source");
  - the returns-export exact-duplicate-row fix (dedupe_returns_export) --
    see its own docstring for why this exists and isn't optional.

Refresh procedure (performed by whoever/whatever is driving the builder
through Claude -- ROADMAP.md §9 "Roles"): for each entry in SOURCES, fetch
the raw bytes via the named connector, pass them to that source's normalize_*
function here, which writes the matching *_SNAPSHOT path. Then run each
builder exactly as it runs today -- no builder code branches on where its
input file came from.
"""
import csv
import json
import os
import sys

import pandas as pd

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_HERE)

# ---------------------------------------------------------------------------
# Pinned sources -- the only place any of these IDs/tabs should live.
# ---------------------------------------------------------------------------

SOURCES = {
    "line_detail": {
        "connector": "google_drive",
        "file_id": "1r5D03e3Df_Qyinrps0wqLlqsp3YGATi6Ob3cvjv7Dok",
        "tab": "Line Detail",
        "note": (
            "Product reference: taxonomy/status/cost. CONFIRMED 2026-08-11: the "
            "live sheet has 2 leading admin rows above the real header (a 'COPY "
            "THIS LINE' row + a blank row) and an extra Barcode column the "
            "committed snapshot doesn't carry -- normalize_line_detail_xlsx "
            "strips the leading rows so the header lands on row 1, matching "
            "what trading/line_detail.py's load_line_detail() assumes."
        ),
    },
    "shopify_product_data": {
        "connector": "google_drive",
        "file_id": "1wb7Xj1ionrL3eoZKBZ5JZTxo6jAiIqFaBQ74IJprcjo",
        "tab": "IN Shopify Product Data",
        "note": (
            "On-hand `Inventory` by `SKU`, for weeks-cover. Lives as a tab "
            "inside the same Yotpo-bound sheet, not its own file (confirmed "
            "2026-08-11 by downloading the sheet and listing tab names -- the "
            "colon in 'IN: Shopify Product Data' is stripped on xlsx export, "
            "so the literal tab name is 'IN Shopify Product Data'). NEVER the "
            "ship_sheet entry below -- that's inbound POs, not on-hand stock."
        ),
    },
    "yotpo_reviews": {
        "connector": "google_drive",
        "file_id": "1wb7Xj1ionrL3eoZKBZ5JZTxo6jAiIqFaBQ74IJprcjo",
        "tab": "API Yotpo Reviews",
        "expected_columns": (
            "ID", "Market", "Title", "Content", "Score", "Votes Up", "Votes Down",
            "Created At", "Updated At", "Sentiment", "SKU", "Name", "Email",
            "Reviewer Type", "Deleted", "Archived", "Escalated", "User Reference",
            "Is Incentivized", "Product SKU", "Product Group", "Product Title",
            "Product Type", "Product Category", "Sub Category", "Collection",
            "Material", "Finish", "Product Type Metafield", "Style Metafield",
            "Descriptor Word", "Product_Type", "YGroup", "Product Status", "CORE?",
            # "Review IDs" deliberately excluded -- see normalize_yotpo_reviews_
            # from_csv's docstring for why it's the one column allowed to be
            # missing entirely, not just blank.
        ),
        "note": (
            "Pulled with deleted=true -- deleted/escalated rows are in this "
            "tab by design, the whole criticism signal review_feedback.py "
            "needs. Same sheet as shopify_product_data above, different tab.\n"
            "CONFIRMED BROKEN 2026-08-13: the sheet's own 'Review IDs' column "
            "(a formula, not raw Yotpo data) reads literal 'Loading...' on row "
            "2 and blank on the other 5,516 of 5,517 rows -- a stuck Google "
            "Sheets formula, not something fixable from here (the Drive "
            "connector reads final values, not formula definitions). Data is "
            "also stale as of this check: newest review 2026-07-05. Refresh "
            "has been deliberately withheld rather than risk undercounting --  "
            "review_feedback.py's dedupe_key falls back to Email+Content when "
            "Review IDs is blank, and that fallback is measurably less precise "
            "at catching syndicated copies (the same review posted once, "
            "attached to every SKU in a product group) than a real grouped ID.\n"
            "ALTERNATE PATH while the Sheet is stuck: normalize_yotpo_reviews_"
            "from_csv() below accepts a raw Yotpo CSV export uploaded directly "
            "(same shape as project knowledge's Sample_of_Yotpo_data_.csv), "
            "bypassing the Sheet (and its broken formula) entirely -- landed "
            "at the exact same snapshot path normalize_yotpo_reviews_xlsx "
            "writes to, so review_feedback.py needs no changes either way."
        ),
    },
    "matrixify_order_slices": {
        "connector": "google_drive",
        "file_id": None,  # TBD -- set once Appendix A's SLICE_FOLDER_ID Script Property exists
        "manifest_file_id": None,  # TBD -- Appendix A's MANIFEST_FILE_ID
        "note": (
            "REPLACES the old matrixify_orders_uk/matrixify_orders_us rolling-"
            "sheet entries (2026-08-13, Matrixify Slice Architecture migration). "
            "Those relied on one ~400-day Drive sheet per store; that sheet's "
            "Google-native -> xlsx export started failing outright (not a size "
            "limit -- the conversion step itself fails for a file this large), "
            "and a 400-day window can never serve a YoY comparison regardless. "
            "Replaced by plain CSV files in a Drive folder, per store per month "
            "(orders_<store>_<YYYY-MM>.csv, written by Appendix A's Apps Script "
            "-- runDaily for the standing pipeline, backfillNext/a local run of "
            "trading/tools/backfill_slice.py for the one-time historical "
            "backfill), plus orders_manifest.csv recording rows/orders/"
            "min_created_at/max_created_at/sha256/last_written per slice -- see "
            "matrixify_orders_snapshot()/load_orders_manifest() below.\n"
            "file_id/manifest_file_id are TBD -- they don't exist yet (Appendix "
            "A's SLICE_FOLDER_ID/MANIFEST_FILE_ID Script Properties haven't been "
            "created). Left honestly unset rather than filled with a placeholder "
            "that looks real. Local landing paths (matrixify_orders_snapshot()/ "
            "ORDERS_MANIFEST_PATH below) work today independent of these Drive "
            "IDs, since trading/tools/backfill_slice.py writes there directly -- "
            "note the LOCAL filename convention (orders_<period>_<STORE>.csv) "
            "is a different naming layer from the Drive one above, not a typo."
        ),
    },
    "returns_zap": {
        "connector": "google_drive",
        "file_id": "1tyinVS7suxKIdaY9Y3R6geaOFkYgeaeVWcDU2VHaGzs",
        "tab": "API: Returns",
        "expected_columns": (
            "Country", "Order Id", "Order Number", "Order Date", "RMA Number",
            "Return Date", "Stage", "Status", "SKU", "Shopify Variant Id",
            "Product", "Variant", "Quantity", "Value", "Return Type", "Return Reason",
        ),
        "note": (
            "ReturnZap pull via the getReturns Apps Script function, key in "
            "Script Properties (Yotpo-style secret boundary). ONE tab holds "
            "BOTH stores (US + UK) -- Country is ship-to (UK/US/ROW), Value "
            "is the line refund value. Tab is named 'API: Returns' in the "
            "Sheets UI, but CONFIRMED 2026-08-12: the colon is stripped on "
            "xlsx export (same quirk as shopify_product_data below) -- the "
            "literal tab name after download is 'API Returns', which is "
            "what normalize_returns_zap_xlsx actually looks for. WHEN THIS "
            "SOURCE MOVES: only the "
            "file_id line above changes -- tab/expected_columns stay pinned "
            "here too, so a rename elsewhere is caught by preflight, not "
            "silently followed. KNOWN BUGS, not fixed here (no Apps Script "
            "connector available to this build -- flag at the source):\n"
            "  (1) 2026-08-11 initial pull: exact full-row duplicates "
            "(1,755 raw / 450 distinct, one row x139) -- almost certainly "
            "appending instead of clearing-then-writing. dedupe_returns_"
            "export() fixes this on read, every time, regardless of scale.\n"
            "  (2) RESOLVED, confirmed 2026-08-13: (2) used to read \"GB/UK "
            "stops dead at Return Date 2025-03-03 and never resumes\" -- as of "
            "today both markets are current (GB 45,838 rows through "
            "2026-08-11, US 8,993 rows through 2026-08-11, 74,218 total raw "
            "rows). Whatever store-specific outage caused the old gap is no "
            "longer present. Left here rather than deleted so a future "
            "session doesn't waste time re-diagnosing something already "
            "checked -- if a similar stall reappears, check per-country "
            "freshness the same way (group raw rows by Country, take max "
            "Return Date per group) before assuming it's the same root cause."
        ),
    },
    "ship_sheet": {
        "connector": "google_drive",
        "file_id": "1wErCQzW2pi8-OnzN2bbuu5bNkTFt1PGhmBumEFzQMaM",
        "tab": "IP Import",
        "note": (
            "INBOUND PURCHASE ORDERS ONLY (Ordered/Received/Remaining per PO, "
            "vendor, cost price) -- stock on order/arriving, never on-hand. "
            "Not read by any fetch function here; kept in the registry only "
            "so nobody re-discovers it and wires it into weeks-cover by "
            "mistake. Use shopify_product_data for on-hand inventory."
        ),
    },
    "revenue_tracker": {
        "connector": "google_drive",
        "file_id": "1gFdzyNvA1IBN1Iy0hwjpf0sXeNHNA-PCudFTXLhkpQ0",
        "tab": "ACT DAY",
        "kind": "reference",
        "note": (
            "REFERENCE, NOT A FEED (item 8, 2026-08-12) -- read-only, used only "
            "to compute an on-dashboard reconciliation LINE tying dashboard "
            "Total Revenue to this tracker's own Gross Revenue figure "
            "(tracker_gross - tracker_shipping ~= dashboard_total, shipping is "
            "the only allowed difference). Never a source any builder computes "
            "its own headline FROM -- the internal country reconciliation "
            "(uk+us+row==total) stays the real integrity check.\n"
            "STILL OPEN, do not wire the actual tie-out until confirmed with "
            "Annie (she offered a walkthrough -- lock the precise cell/"
            "definition with her, don't infer it):\n"
            "  - Exact source figure: candidate found 2026-08-12 by reading the "
            "sheet directly -- 'ACT DAY' has a 'Gross Revenue' block (D2C: "
            "rows 139-177, its own UK/US/Other sub-rows + a 'TOTAL' row at "
            "167; B2B: rows 297+, analogous shape) with each region's own "
            "'Shipping' sub-row already broken out -- but which exact row/sum "
            "Annie means by 'Gross Revenue' hasn't been confirmed with her.\n"
            "  - VAT basis: the sheet has a SEPARATE 'Incl VAT' row (ACT DAY "
            "row 510) used only for the ROAS efficiency metric, distinct from "
            "the 'Gross Revenue' block above -- circumstantial evidence Gross "
            "Revenue itself is ex-VAT (matching this dashboard's own basis), "
            "but not confirmed with Annie.\n"
            "  - Region mapping: tracker rows are literally 'UK'/'US'/'Other' "
            "-- Other -> ROW mapping is very likely direct but unconfirmed."
        ),
    },
    "matrixify_uk": {
        "connector": "mcp",
        "server": "Matrixify-PlankUK",
        "note": "UK store order exports (Orders entity, filtered per period).",
    },
    "matrixify_us": {
        "connector": "mcp",
        "server": "Matrixify-PlankUS",
        "note": "US store order exports (Orders entity, filtered per period).",
    },
}

# Where a fetched snapshot lands -- every builder reads from these paths,
# whether freshly refreshed or a few days stale. trading/source/*.xlsx here
# is a rolling product-reference snapshot (refreshed every run), NOT a
# frozen monthly export -- CLAUDE.md's "never overwrite a committed month's
# export" exception is about the Matrixify order CSVs, not this file.
SNAPSHOT_DIR = os.path.join(_REPO_ROOT, "source")
TRADING_SNAPSHOT_DIR = os.path.join(_REPO_ROOT, "trading", "source")

LINE_DETAIL_SNAPSHOT = os.path.join(TRADING_SNAPSHOT_DIR, "line_detail.xlsx")
SHOPIFY_INVENTORY_SNAPSHOT = os.path.join(TRADING_SNAPSHOT_DIR, "shopify_inventory.csv")
RETURNS_ZAP_SNAPSHOT = os.path.join(SNAPSHOT_DIR, "returns_zap.csv")
YOTPO_REVIEWS_SNAPSHOT = os.path.join(SNAPSHOT_DIR, "yotpo_reviews.csv")

# Matrixify orders -- per-store, per-month CSV slices + a manifest (2026-08-13
# migration, superseding the single rolling orders_ALL_<STORE>.csv this
# module used until then). Written by trading/tools/backfill_slice.py (the
# one-time local backfill) and, eventually, by the standing Apps Script in
# Appendix A of the Matrixify Slice Architecture brief -- one shared
# manifest regardless of which produced a given slice.
ORDERS_MANIFEST_PATH = os.path.join(TRADING_SNAPSHOT_DIR, "orders_manifest.csv")
ORDERS_MANIFEST_COLS = ("store", "period", "file_name", "rows", "orders",
                        "min_created_at", "max_created_at", "sha256", "last_written")


def matrixify_orders_snapshot(store, period):
    """trading/source/orders_<period>_<STORE>.csv -- e.g.
    "trading/source/orders_2025-07_UK.csv" -- the per-store, per-month slice
    for `period` ("2026-06"). This is the LOCAL landing-path convention
    (period first, STORE uppercase, no subdirectory) -- a different naming
    layer from the Drive-side per-slice filename in the brief's own §2.1
    (orders_<store>_<YYYY-MM>.csv, lowercase store); don't conflate the two,
    a mismatch here means the builders find nothing on disk.

    `period` is now REQUIRED: it used to be accepted-and-ignored, back when
    one rolling ~400-day file served every period unfiltered (retired
    2026-08-13 -- that file's own Drive sheet stopped being downloadable at
    all, and no 400-day window could ever serve a YoY comparison regardless
    of size). There is no longer a single file that could serve an
    arbitrary period, by design: matrixify_orders_snapshot_covers() below
    needs a specific period to check coverage FOR. Never committed
    (source/ is gitignored, no exceptions, same as every other snapshot
    here).
    """
    return os.path.join(TRADING_SNAPSHOT_DIR, f"orders_{period}_{store.upper()}.csv")


def load_orders_manifest(manifest_path=ORDERS_MANIFEST_PATH):
    """{(store, period): row_dict} from orders_manifest.csv, {} if the
    manifest doesn't exist yet (e.g. before any backfill/standing slice has
    ever been written). Keys are the lowercase store/period strings exactly
    as trading/tools/backfill_slice.py and the Appendix A Apps Script both
    write them.
    """
    if not os.path.exists(manifest_path):
        return {}
    with open(manifest_path, newline="", encoding="utf-8") as fh:
        return {(row["store"], row["period"]): row for row in csv.DictReader(fh)}


def matrixify_orders_snapshot_covers(store, period, manifest=None):
    """Full-period coverage from the manifest, not "at least one row" --
    the bug this replaces (the old matrixify_orders_snapshot_covers(csv_path,
    period), which scanned a CSV and returned True on the FIRST matching
    row, then got OR'd across stores at both call sites) let a missing store
    pass silently, since a rolling file merely existing was never proof it
    covered a given month either.

    A manifest row's min_created_at/max_created_at must bracket the whole
    period (on/before its first day, on/after its last), not just intersect
    it -- a slice that only half-landed (an interrupted backfill, say) must
    not read as covered. Caveat, intentional rather than fixed further: a
    real calendar day with genuinely zero orders at the very start or end of
    a period would also fail this check, same as a missing day would --
    implausible at this store's volume, and a false negative here is a far
    smaller risk than the "any single row passes" bug it replaces.
    """
    manifest = manifest if manifest is not None else load_orders_manifest()
    row = manifest.get((store.lower(), period))
    if row is None or not row.get("min_created_at") or not row.get("max_created_at"):
        return False

    from datetime import datetime
    from zoneinfo import ZoneInfo
    from common.period import month_key_bounds

    london = ZoneInfo("Europe/London")
    period_start, period_end = month_key_bounds(period)
    min_dt = datetime.strptime(row["min_created_at"], "%Y-%m-%d %H:%M:%S %z").astimezone(london).date()
    max_dt = datetime.strptime(row["max_created_at"], "%Y-%m-%d %H:%M:%S %z").astimezone(london).date()
    return min_dt <= period_start and max_dt >= period_end


def assert_orders_coverage(period, stores=("uk", "us")):
    """The shared, AND-across-stores coverage guard (Matrixify Slice
    Architecture brief 5/11.1) -- both trading/build_matrixify_dashboard.py
    (which had an OR-across-stores bug) and returns/build_dashboard.py
    (which had NO coverage guard at all) call this, so there is one place
    deciding "is this period's Matrixify data actually usable," not a
    duplicated or drifted copy per caller. Lives in common/, not trading/,
    so returns/build_dashboard.py can call it without depending on
    trading/. Raises AssertionError naming the exact missing/partial
    store(s), never returns a soft False a caller could accidentally OR past.
    """
    manifest = load_orders_manifest()
    missing = [s for s in stores if not matrixify_orders_snapshot_covers(s, period, manifest=manifest)]
    assert not missing, (
        f"ORDERS COVERAGE FAILED: {period} is not fully covered for store(s) {missing} "
        f"-- a missing/partial store must not silently pass (this used to be OR'd "
        f"across stores; that's exactly the bug this replaces). Check "
        f"{ORDERS_MANIFEST_PATH} has a row for each store, and that its "
        f"min_created_at/max_created_at actually bracket the whole period."
    )


def check_orders_manifest_staleness(stores=("uk", "us"), max_staleness_hours=36, as_of=None):
    """WARNS (does not abort) if any store's newest manifest last_written is
    older than max_staleness_hours -- the standing Apps Script's runDaily
    trigger going silent is otherwise invisible until someone notices a
    number looks wrong (brief section 5's staleness assertion).
    """
    from datetime import datetime, timezone

    as_of = as_of or datetime.now(timezone.utc)
    manifest = load_orders_manifest()
    for store in stores:
        rows = [r for (s, _p), r in manifest.items() if s == store.lower()]
        if not rows:
            print(f"check_orders_manifest_staleness: no manifest rows at all for store {store!r}",
                  file=sys.stderr)
            continue
        newest = max(datetime.fromisoformat(r["last_written"]) for r in rows)
        staleness_hours = (as_of - newest).total_seconds() / 3600
        if staleness_hours > max_staleness_hours:
            print(f"check_orders_manifest_staleness: WARNING -- {store} manifest's newest "
                  f"last_written is {newest.isoformat()} ({staleness_hours:.1f}h before "
                  f"{as_of.isoformat()}) -- the standing Apps Script pipeline may have gone silent.",
                  file=sys.stderr)


# ---------------------------------------------------------------------------
# Normalizers -- turn a raw connector download into the shape the EXISTING
# parsers already expect. Each takes a raw downloaded file/bytes and writes
# the matching *_SNAPSHOT path; none of them change any builder's parser.
# ---------------------------------------------------------------------------

def _clean_cell(v):
    """openpyxl reads whole-number cells back as Python floats (12021886714236.0),
    even though the underlying Shopify/Matrixify value is a true integer (an
    order ID, a line ID, a quantity). Left as-is, that trailing '.0' becomes
    part of the string when written to CSV -- harmless for matrixify_source.py
    (which runs everything through float() anyway), but silently fatal for
    returns/build.py's order-id JOIN: the sales side reads this CSV with plain
    csv.DictReader (no numeric coercion, so "12021886714236.0" stays exactly
    that), while the returns side casts ReturnZap's Order Id through pandas'
    Int64 (which cleans "5654760980759.0" -> "5654760980759") -- two clean-
    looking IDs that never actually match as strings. CONFIRMED 2026-08-12:
    this was found by the returns builder's own overlap gate correctly
    refusing to publish a 0%-overlap join rather than a subtler wrong number.
    """
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def normalize_returns_zap_xlsx(raw_xlsx_path, out_path=RETURNS_ZAP_SNAPSHOT):
    """raw_xlsx_path: the ReturnZap Apps Script's sheet, downloaded as-is
    (single relevant tab, 'API: Returns' in the Sheets UI -- populated by
    the getReturns Apps Script function, see
    claude/returnzap_setup_runbook.md in project knowledge). CONFIRMED
    2026-08-12: the colon is stripped on xlsx export (same quirk already
    documented for 'IN: Shopify Product Data' -> 'IN Shopify Product Data'
    above) -- the literal tab name in a downloaded xlsx is 'API Returns',
    not 'API: Returns'. Extracts that tab to a plain CSV at out_path in the
    sheet's own native column names (Order Id, SKU, Quantity, Stage, ...)
    -- load_returns_zap_snapshot()/load_returns_export_from_sheet() read
    this directly, no parser change needed. openpyxl's None becomes an
    empty string, matching what pd.read_csv would see from a real CSV
    export.
    """
    import openpyxl

    wb = openpyxl.load_workbook(raw_xlsx_path, read_only=True, data_only=True)
    ws = wb["API Returns"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows:
            if row[0] is None and all(v is None for v in row):  # trailing blank row
                continue
            writer.writerow(["" if v is None else _clean_cell(v) for v in row])
    return out_path


def normalize_line_detail_xlsx(raw_xlsx_path, out_path=LINE_DETAIL_SNAPSHOT):
    """raw_xlsx_path: the live Drive sheet, downloaded as-is (2 leading admin
    rows above the real header, confirmed 2026-08-11). Writes a new xlsx at
    out_path with the real header on row 1 and every data row below it,
    unchanged otherwise (Barcode column and all) -- trading/line_detail.py's
    load_line_detail() takes row 1 as header unconditionally, so this is the
    only change needed for the live sheet to parse exactly like the
    committed local snapshot always has.
    """
    import openpyxl

    src = openpyxl.load_workbook(raw_xlsx_path, read_only=True, data_only=True)
    ws = src["Line Detail"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        i for i, row in enumerate(rows)
        if row and any(str(c).strip() == "SKU" for c in row if c is not None)
    )
    out = openpyxl.Workbook()
    out_ws = out.active
    out_ws.title = "Line Detail"
    for row in rows[header_idx:]:
        out_ws.append(row)
    out.save(out_path)
    return out_path


def normalize_shopify_inventory_xlsx(raw_xlsx_path, out_path=SHOPIFY_INVENTORY_SNAPSHOT):
    """raw_xlsx_path: the Yotpo-bound sheet downloaded as xlsx. Extracts the
    `IN Shopify Product Data` tab's SKU/Inventory columns to a flat CSV --
    trading only needs on-hand units by SKU, not the other ~30 product-feed
    columns in that tab.
    """
    import openpyxl

    wb = openpyxl.load_workbook(raw_xlsx_path, read_only=True, data_only=True)
    ws = wb["IN Shopify Product Data"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    sku_idx = header.index("SKU")
    inv_idx = header.index("Inventory")
    df = pd.DataFrame(
        [(r[sku_idx], r[inv_idx]) for r in rows if r[sku_idx]],
        columns=["SKU", "Inventory"],
    )
    df.to_csv(out_path, index=False)
    return out_path


def normalize_yotpo_reviews_xlsx(raw_xlsx_path, out_path=YOTPO_REVIEWS_SNAPSHOT):
    """raw_xlsx_path: the Yotpo-bound sheet downloaded as xlsx. Extracts the
    `API Yotpo Reviews` tab to a flat CSV.

    Found 2026-08-12 refreshing this snapshot for the first time since the
    committed one: openpyxl reads a whole-number Google Sheets cell (ID,
    Score, Votes Up/Down -- confirmed the affected columns) back as a Python
    float, so a naive dump renders "866558445" as "866558445.0". Confirmed
    by diffing a fresh pull against the committed snapshot: every row
    "changed" until this was stripped, at which point the two were
    byte-identical except for 2 genuinely new reviews. Cast every
    whole-number float to int before writing so this doesn't recur on every
    future refresh.
    """
    import openpyxl

    wb = openpyxl.load_workbook(raw_xlsx_path, read_only=True, data_only=True)
    ws = wb["API Yotpo Reviews"]

    def _clean(v):
        if isinstance(v, float) and v.is_integer():
            return int(v)
        return v

    raw_rows = [[_clean(v) for v in row] for row in ws.iter_rows(values_only=True)]
    # Trailing blank column/rows (found 2026-08-12): ws.max_row/max_column on
    # the live sheet run past the real data by a formatting artifact -- an
    # all-blank trailing header cell and 2 fully-blank trailing rows, neither
    # present in the committed snapshot. Confirmed via diff against it that
    # trimming these (not the real header width) reproduces it exactly bar
    # genuinely new reviews.
    header = raw_rows[0]
    width = max(i + 1 for i, v in enumerate(header) if v not in (None, ""))
    rows = [row[:width] for row in raw_rows if any(v not in (None, "") for v in row[:width])]
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerows(rows)
    return out_path


def normalize_yotpo_reviews_from_csv(raw_csv_path, out_path=YOTPO_REVIEWS_SNAPSHOT):
    """Alternate path for when the Yotpo Google Sheet's own 'Review IDs'
    formula is broken/stuck (see yotpo_reviews's SOURCES note) -- refreshing
    from the Sheet while that's stuck would silently undercount syndicated
    reviews via review_feedback.py's own fallback dedupe key (Email +
    Content), which is measurably less precise than a real grouped ID.
    Accepts a raw Yotpo CSV export (the same shape as project knowledge's
    Sample_of_Yotpo_data_.csv) uploaded directly instead -- e.g. someone
    attaches the file to a chat and hands its /mnt/user-data/uploads/ path
    here -- bypassing the Sheet (and its broken formula) entirely.

    Deliberately tolerant of a missing 'Review IDs' column altogether, not
    just a blank one: that column looks like something computed inside
    Plank's own Google Sheet, not part of Yotpo's native export -- a raw
    pull straight from Yotpo's own UI may never have had it. Either way,
    review_feedback.py's existing fallback already handles a blank/absent
    Review IDs value per row; this function doesn't try to compute one
    itself. Its only job is landing the raw rows at the exact CSV shape
    normalize_yotpo_reviews_xlsx already produces, so review_feedback.py
    (or anything else downstream) can't tell which path a given refresh
    came from -- same snapshot path, same column names, same dtype
    handling it would have gotten from the Sheet.
    """
    with open(raw_csv_path, newline="", encoding="utf-8-sig") as f:
        rows = [row for row in csv.reader(f) if any(v.strip() for v in row)]

    if not rows:
        raise ValueError(f"normalize_yotpo_reviews_from_csv: {raw_csv_path} has no data rows")

    header = rows[0]
    expected = SOURCES["yotpo_reviews"]["expected_columns"]
    missing = [c for c in expected if c not in header]
    if missing:
        raise ValueError(
            f"normalize_yotpo_reviews_from_csv: {raw_csv_path} is missing expected column(s) "
            f"{missing} -- doesn't look like the Yotpo export format review_feedback.py expects. "
            f"Check it's a genuine Yotpo reviews export, not a different report."
        )
    if "Review IDs" not in header:
        print("normalize_yotpo_reviews_from_csv: no 'Review IDs' column in this export -- "
              "review_feedback.py's fallback dedupe key (Email + Content) will be used for "
              "every row, same degraded-but-working behavior as when the Sheet's formula is "
              "stuck. Syndicated-review counts may be very slightly less precise as a result.",
              file=sys.stderr)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerows(rows)
    return out_path


def dedupe_returns_export(df):
    """Drop exact full-row duplicates from a raw ReturnZap pull before
    anything downstream sees it.

    Found 2026-08-11: the live sheet had 1,755 raw rows but only 450
    distinct -- one return duplicated 139 times, with Stage/Status/RMA
    Number/Return Date byte-identical across every copy. This is NOT
    stage-history (genuine multi-RMA returns on the same order+SKU do exist
    and legitimately differ by RMA Number/Return Date -- confirmed
    separately those are untouched by this function; only true full-row
    copies are dropped). Left unfixed, returns/build.py's existing
    sku+order dedupe (`.agg(qty=("qty", "sum"))`) would SUM across the
    duplicate copies too, overstating returned units/cash by ~4x on average
    and up to 139x for specific rows -- almost certainly the getReturns
    Apps Script appending on every run instead of clearing-then-writing
    (the documented Yotpo pattern). Returns (deduped_df, n_dropped) -- the
    caller must report n_dropped, never silently absorb it (see
    common/reconciliation_gate.assert_bucket_reported).
    """
    before = len(df)
    deduped = df.drop_duplicates().reset_index(drop=True)
    dropped_exact = before - len(deduped)

    # Second pass, found 2026-08-12 re-verifying order- vs unit-level
    # counting after the full-history pull (74,218 raw rows): a small
    # residual pattern survives the exact-row pass -- the SAME line (same
    # Order Id + SKU + RMA Number + Shopify Variant Id + Quantity)
    # re-captured moments apart with exactly one other field updated
    # (Return Date bumped a few seconds, or Stage/Status/Value/Return
    # Reason changed) -- a state-snapshot artifact, not a second real
    # return. Confirmed on 10 of 15,366 post-exact-dedupe rows (2026-08-12):
    # every affected order already has >=1 OTHER row for that order, so
    # returns/build.py's order-level count (`ret["order"].nunique()`) is
    # structurally unaffected either way -- nunique() on order already
    # collapses duplicate rows for the SAME order to one, regardless of
    # how many SKU-lines or duplicate-lines exist within it (verified with
    # a synthetic 5x-duplicate case). This pass only prevents units_returned
    # (a secondary, non-headline metric per ROADMAP.md's locked "orders-
    # based rate" decision) from double-counting the same physical unit --
    # keeps the row with the latest Return Date (the most up-to-date
    # snapshot) per (Order Id, SKU, RMA Number, Shopify Variant Id, Quantity).
    key_cols = ["Order Id", "SKU", "RMA Number", "Shopify Variant Id", "Quantity"]
    if all(c in deduped.columns for c in key_cols) and "Return Date" in deduped.columns:
        sort_col = pd.to_datetime(deduped["Return Date"], utc=True, errors="coerce")
        deduped = (
            deduped.assign(_sort=sort_col)
            .sort_values("_sort")
            .drop_duplicates(subset=key_cols, keep="last")
            .drop(columns="_sort")
            .sort_index()
            .reset_index(drop=True)
        )
    dropped_snapshot = before - dropped_exact - len(deduped)
    dropped = before - len(deduped)
    if dropped_snapshot:
        print(f"dedupe_returns_export: also collapsed {dropped_snapshot} same-line "
              f"snapshot-update row(s) (same order+SKU+RMA+variant+quantity, one field "
              f"changed between captures) -- kept the latest Return Date per line",
              file=__import__("sys").stderr)
    return deduped, dropped


def load_shopify_inventory(csv_path=SHOPIFY_INVENTORY_SNAPSHOT):
    """-> {sku: on_hand_units}, from the landed shopify_product_data
    snapshot. Never sourced from ship_sheet -- see its SOURCES note.
    """
    df = pd.read_csv(csv_path)
    df["SKU"] = df["SKU"].astype(str).str.strip()
    df = df[df["SKU"].notna() & (df["SKU"] != "") & (df["SKU"].str.lower() != "nan")]
    inv = pd.to_numeric(df["Inventory"], errors="coerce").fillna(0)
    return dict(zip(df["SKU"], inv))


def write_snapshot_metadata(csv_path, file_id, source_modified_time, fetched_at):
    """Record which Drive file/version a landed snapshot came from --
    file_id + the source's own last-modified time, alongside when this
    fetch happened (caller-supplied, real wall-clock time -- this module
    never calls datetime.now() itself). Lets a future run tell whether the
    live sheet has changed since this snapshot was taken without
    re-downloading to check. Sidecar file, `<csv_path>.meta.json` --
    doesn't change how any reader loads the CSV itself.
    """
    meta = {
        "file_id": file_id,
        "source_modified_time": source_modified_time,
        "fetched_at": fetched_at,
    }
    meta_path = csv_path + ".meta.json"
    with open(meta_path, "w") as f:
        json.dump(meta, f, indent=2)
    return meta_path


def check_yotpo_freshness(csv_path=YOTPO_REVIEWS_SNAPSHOT, as_of=None, max_staleness_days=14):
    """Part 4 (period-from-prompt, 2026-08-12): Yotpo reviews follow the
    REGION filter only, per the returns brief -- never the trading/returns
    period. review_feedback.py already applies no period filter (confirmed
    2026-08-12), so there is nothing to slice here; this just confirms the
    snapshot itself is present and recent, the same freshness pattern as
    returns/validate.py's check_freshness. Warns (does not abort) if the
    snapshot's newest `Created At` is older than max_staleness_days.
    """
    from datetime import date
    import sys as _sys

    if as_of is None:
        as_of = date.today()
    df = pd.read_csv(csv_path)
    dates = pd.to_datetime(df["Created At"], format="%d-%b-%Y", errors="coerce").dt.date.dropna()
    newest = dates.max()
    staleness = (as_of - newest).days
    if staleness > max_staleness_days:
        print(f"check_yotpo_freshness: WARNING -- newest review Created At is {newest} "
              f"({staleness} days before {as_of}) -- the Yotpo pull may not have run recently.",
              file=_sys.stderr)
    return newest, staleness


def load_returns_zap_snapshot(csv_path=RETURNS_ZAP_SNAPSHOT):
    """Read the landed ReturnZap snapshot and apply dedupe_returns_export.
    Returns (df, n_dropped) in the sheet's own native column names (Order
    Id, SKU, Quantity, Stage, ...) -- returns/build.py's
    load_returns_export_from_sheet() maps these to the standardized shape
    the rest of the pipeline already consumes.
    """
    df = pd.read_csv(csv_path, encoding="utf-8-sig", low_memory=False)
    return dedupe_returns_export(df)


# ---------------------------------------------------------------------------
# Preflight -- validates already-landed snapshots (shape/columns/non-empty).
# Does NOT check live connector reachability itself (see module docstring:
# that requires the Drive/Matrixify MCP tools, only reachable from an
# interactive Claude session) -- run this after a refresh to confirm what
# landed is actually usable, naming the exact source on failure.
# ---------------------------------------------------------------------------

_EXPECTED_COLUMNS = {
    LINE_DETAIL_SNAPSHOT: None,  # checked via load_line_detail() itself (header lookup by name)
    SHOPIFY_INVENTORY_SNAPSHOT: {"SKU", "Inventory"},
    # Derived from SOURCES["returns_zap"] rather than re-listed -- one place
    # to update if the sheet's columns ever change.
    RETURNS_ZAP_SNAPSHOT: set(SOURCES["returns_zap"]["expected_columns"]),
}


def preflight_returns_zap(csv_path=RETURNS_ZAP_SNAPSHOT):
    """Returns-connection-scoped preflight: fails loud naming the exact
    problem, never compares against a prior report (that's what the
    reconciliation gate/regression fixture do, not this). Checks:
      - the snapshot loads with every expected column present, non-empty;
      - the RMA Number + SKU duplicate count is logged (not silently
        absorbed) -- there is no separate "line" id in this feed, so
        RMA+SKU is the finest natural key it actually carries; a remaining
        duplicate there is a genuine multi-line return, not noise (do not
        confuse with the exact-full-row dedupe, a stricter, separate check);
      - BOTH markets are present -- non-zero US and UK/GB row counts,
        catching a missing store key outright.

    Also logs (informational -- a staleness signal, not a shape one) the
    most recent Order Date per market. Found 2026-08-11: the UK pull
    stopped dead at Return Date 2025-03-03 while US continued to
    2026-08-01 -- a bare non-zero-rows check does NOT catch this (UK rows
    exist, just none recent), so this is logged every run to make a future
    market-specific staleness visible without re-deriving it by hand.
    """
    if not os.path.exists(csv_path):
        raise AssertionError(f"PREFLIGHT FAILED: returns_zap -- missing snapshot at {csv_path}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", low_memory=False)
    expected_cols = set(SOURCES["returns_zap"]["expected_columns"])
    missing_cols = expected_cols - set(df.columns)
    problems = []
    if missing_cols:
        problems.append(f"returns_zap: missing column(s) {sorted(missing_cols)} in {csv_path}")
    if df.empty:
        problems.append(f"returns_zap: snapshot at {csv_path} is empty")
    if problems:
        raise AssertionError("PREFLIGHT FAILED:\n  " + "\n  ".join(problems))

    deduped, n_full_dupes = dedupe_returns_export(df)
    n_rma_sku_dupes = len(deduped) - len(deduped.drop_duplicates(subset=["RMA Number", "SKU"]))
    print(f"preflight[returns_zap]: {len(df)} raw rows -> {len(deduped)} after exact-duplicate-row "
          f"dedupe ({n_full_dupes} dropped); {n_rma_sku_dupes} row(s) still share an RMA Number+SKU "
          f"key after that (genuine multi-line returns, not noise -- returns/build.py's own gate is "
          f"sku+order, the ROADMAP.md-locked key, not RMA+SKU+line).")

    counts = deduped["Country"].value_counts()
    us_rows = int(counts.get("US", 0))
    uk_rows = int(counts.get("GB", 0) + counts.get("UK", 0))
    if us_rows == 0:
        problems.append("returns_zap: zero US rows in the snapshot -- a market is silently missing")
    if uk_rows == 0:
        problems.append("returns_zap: zero UK/GB rows in the snapshot -- a market is silently missing")
    if problems:
        raise AssertionError("PREFLIGHT FAILED:\n  " + "\n  ".join(problems))
    print(f"preflight[returns_zap]: both markets present (US={us_rows} rows, UK={uk_rows} rows)")

    order_date = pd.to_datetime(deduped["Order Date"], utc=True, errors="coerce")
    for label in ("US", "GB"):
        sub_dates = order_date[deduped["Country"] == label]
        if len(sub_dates):
            print(f"preflight[returns_zap]: most recent {label} Order Date in the snapshot: "
                  f"{sub_dates.max()} -- a market with only OLD rows passes the non-zero check "
                  f"above but is effectively stale; check this before trusting a build for a "
                  f"recent period.")
    return {"us_rows": us_rows, "uk_rows": uk_rows, "exact_duplicates_dropped": n_full_dupes}


def preflight():
    """Assert every landed snapshot exists, is non-empty, and has the
    columns its reader expects -- fails loud, naming the exact source.
    """
    problems = []

    if not os.path.exists(LINE_DETAIL_SNAPSHOT):
        problems.append(f"line_detail: missing snapshot at {LINE_DETAIL_SNAPSHOT}")
    else:
        import openpyxl
        wb = openpyxl.load_workbook(LINE_DETAIL_SNAPSHOT, read_only=True, data_only=True)
        if "Line Detail" not in wb.sheetnames:
            problems.append(f"line_detail: no 'Line Detail' sheet in {LINE_DETAIL_SNAPSHOT}")
        else:
            header = next(wb["Line Detail"].iter_rows(values_only=True))
            if "SKU" not in header:
                problems.append(
                    f"line_detail: header row has no SKU column ({LINE_DETAIL_SNAPSHOT}) -- "
                    f"was normalize_line_detail_xlsx run on the raw download?"
                )

    for path, expected_cols in ((SHOPIFY_INVENTORY_SNAPSHOT, _EXPECTED_COLUMNS[SHOPIFY_INVENTORY_SNAPSHOT]),
                                (RETURNS_ZAP_SNAPSHOT, _EXPECTED_COLUMNS[RETURNS_ZAP_SNAPSHOT])):
        if not os.path.exists(path):
            problems.append(f"{os.path.basename(path)}: missing snapshot at {path}")
            continue
        df = pd.read_csv(path, encoding="utf-8-sig", nrows=5)
        missing = expected_cols - set(df.columns)
        if missing:
            problems.append(f"{os.path.basename(path)}: missing column(s) {sorted(missing)} in {path}")
        if os.path.getsize(path) < 10:
            problems.append(f"{os.path.basename(path)}: empty file at {path}")

    if os.path.exists(RETURNS_ZAP_SNAPSHOT):
        try:
            preflight_returns_zap()
        except AssertionError as e:
            problems.append(str(e))

    if problems:
        raise AssertionError("PREFLIGHT FAILED:\n  " + "\n  ".join(problems))
    print("preflight: all landed snapshots present and shaped as expected")


if __name__ == "__main__":
    preflight()
