"""Tests for common/sku_cuts.py and the By-SKU companion tab it feeds.

Run: python trading/tests/test_sku_cuts.py

These are unit tests over synthetic lines on purpose: the point is to prove
the routing and the arithmetic, which a real month can't isolate. End-to-end
verification against real order lines is a separate step and needs the
period's Matrixify export staged.
"""
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO = os.path.join(_HERE, "..", "..")
for _p in (_REPO, os.path.join(_REPO, "trading"), os.path.join(_REPO, "trading", "dashboard")):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from common import sku_cuts

FAILURES = []


def check(label, got, want, tol=1e-9):
    ok = (got is None and want is None) or (
        got is not None and want is not None and abs(got - want) <= tol)
    print(f"  {'PASS' if ok else 'FAIL'}  {label}: got {got!r}, want {want!r}")
    if not ok:
        FAILURES.append(label)


def test_partition():
    """Every line lands in total, its country, its channel and its cross --
    and nowhere else. This is what assert_cuts_reconcile guards in the build.
    """
    print("test_partition")
    c = sku_cuts.new_cuts()
    # UK D2C: 100 rev, 10 units, cost 2/unit
    sku_cuts.add_line(c, 100.0, 10, "D2C", "UK", 2.0)
    # UK B2B: 60 rev, 12 units
    sku_cuts.add_line(c, 60.0, 12, "B2B", "UK", 2.0)
    # US D2C: 40 rev, 5 units
    sku_cuts.add_line(c, 40.0, 5, "D2C", "US", 2.0)
    # ROW B2B: 20 rev, 3 units
    sku_cuts.add_line(c, 20.0, 3, "B2B", "ROW", 2.0)

    check("total rev", c["total"]["rev"], 220.0)
    check("total units", c["total"]["u"], 30)
    check("uk rev", c["uk"]["rev"], 160.0)
    check("us rev", c["us"]["rev"], 40.0)
    check("row rev", c["row"]["rev"], 20.0)
    check("d2c rev", c["d2c"]["rev"], 140.0)
    check("b2b rev", c["b2b"]["rev"], 80.0)
    check("uk_d2c rev", c["uk_d2c"]["rev"], 100.0)
    check("uk_b2b rev", c["uk_b2b"]["rev"], 60.0)
    check("us_b2b rev (no such line)", c["us_b2b"]["rev"], 0.0)
    check("country cuts sum to total", sum(c[k]["rev"] for k in sku_cuts.COUNTRY_CUTS), 220.0)
    check("cross cuts sum to total", sum(c[k]["rev"] for k in sku_cuts.CROSS_CUTS), 220.0)
    sku_cuts.assert_cuts_reconcile("TEST-SKU", c)
    print("  PASS  assert_cuts_reconcile accepted a well-formed partition")


def test_realised_gm_differs_by_channel():
    """The whole reason GM is computed per cut rather than copied from the
    catalogue: a discounted channel must show a thinner realised margin.
    D2C sells 10 units at 10 each (100); B2B sells 10 units at 6 each (60);
    cost is 2/unit in both. So D2C GM = 1 - 20/100 = 80%, B2B = 1 - 20/60 =
    66.7%. A catalogue gm_pct would have printed the same number twice.
    """
    print("test_realised_gm_differs_by_channel")
    c = sku_cuts.new_cuts()
    sku_cuts.add_line(c, 100.0, 10, "D2C", "UK", 2.0)
    sku_cuts.add_line(c, 60.0, 10, "B2B", "UK", 2.0)
    check("d2c gm", sku_cuts.gm_of(c["d2c"]), 0.80)
    check("b2b gm", sku_cuts.gm_of(c["b2b"]), 1 - 20.0 / 60.0)
    check("total gm", sku_cuts.gm_of(c["total"]), 1 - 40.0 / 160.0)


def test_unknown_cost_yields_none_not_zero():
    """A SKU missing a supplier cost must produce gm=None, never 0.0 (which
    would read as a real, catastrophic margin) and never a fabricated one.
    """
    print("test_unknown_cost_yields_none_not_zero")
    c = sku_cuts.new_cuts()
    sku_cuts.add_line(c, 100.0, 10, "D2C", "UK", None)
    check("cost stays None", c["total"]["cost"], None)
    check("gm is None", sku_cuts.gm_of(c["total"]), None)
    check("rev still counted", c["total"]["rev"], 100.0)


def test_negative_and_zero_denominators():
    """A net-negative cut (all returns, no sales) has no meaningful margin or
    share, and an empty denominator yields None rather than a divide-by-zero
    or a misleading 0.0%.
    """
    print("test_negative_and_zero_denominators")
    c = sku_cuts.new_cuts()
    sku_cuts.add_line(c, -50.0, 0, "D2C", "UK", 2.0)
    check("negative rev gm", sku_cuts.gm_of(c["total"]), None)
    check("share of empty denom", sku_cuts.share_of(c["uk"], c["us"]), None)


def test_roundtrip():
    """serialize -> deserialize must preserve every cut, and a payload missing
    a cut must degrade to zero rather than KeyError.
    """
    print("test_roundtrip")
    c = sku_cuts.new_cuts()
    sku_cuts.add_line(c, 12.345678, 3, "B2B", "US", 1.5)
    back = sku_cuts.deserialize(sku_cuts.serialize(c))
    check("roundtrip rev", back["us_b2b"]["rev"], 12.345678, tol=1e-6)
    check("roundtrip units", back["us_b2b"]["u"], 3)
    partial = sku_cuts.deserialize({"total": {"rev": 5.0, "u": 1, "cost": None}})
    check("missing cut zeroed", partial["uk_d2c"]["rev"], 0.0)


def test_reconcile_gate_catches_a_leak():
    """The gate must actually fail on a leak -- a cut hand-edited so the
    country cuts no longer sum to total.
    """
    print("test_reconcile_gate_catches_a_leak")
    c = sku_cuts.new_cuts()
    sku_cuts.add_line(c, 100.0, 10, "D2C", "UK", 2.0)
    c["uk"]["rev"] = 90.0  # simulate a routing bug
    try:
        sku_cuts.assert_cuts_reconcile("LEAKY-SKU", c)
    except AssertionError as e:
        print(f"  PASS  gate fired: {e}")
        return
    print("  FAIL  gate did not fire on a 10% leak")
    FAILURES.append("reconcile gate")


def test_by_sku_tab_writes_and_ties():
    """Write the real companion tab from a small hand-built contract and read
    the cells back: header/format/block counts line up, and the row's cut
    columns tie to its total.
    """
    print("test_by_sku_tab_writes_and_ties")
    from openpyxl import Workbook, load_workbook
    from excel_companion import _build_by_sku, _SKU_HEADERS

    c = sku_cuts.new_cuts()
    sku_cuts.add_line(c, 100.0, 10, "D2C", "UK", 2.0)
    sku_cuts.add_line(c, 60.0, 12, "B2B", "US", 2.0)
    sku_cuts.add_line(c, 20.0, 2, "D2C", "ROW", 2.0)

    contract = {
        "contract_version": "1.1",
        "current": {"total_sales": 180.0, "units": 24, "d2c_gbp": 120.0, "b2b_gbp": 60.0,
                     "uk_gbp": 100.0, "us_gbp": 60.0, "row_gbp": 20.0, "gm_pct": 0.75},
        "skus_all": [{
            "sku": "TEST-001", "desc": "Test Handle", "coll": "TESTER", "type_": "Cabinetry",
            "item_type": "Handle", "style": "T Bar", "material": "Brass", "finish": "Antique Brass",
            "uk_status": "Live", "us_status": "Live", "gross": 180.0, "units": 24,
            "vslq": 0.15, "gm": 0.72, "inv": 100, "st": 0.19, "wc": 18.0,
            "d2c": 120.0, "b2b": 60.0, "uk": 100.0, "uk_u": 10, "us": 60.0, "us_u": 12,
            "row": 20.0, "row_u": 2, "is_kit": False, "supplier_cost": 2.0,
            "lq": 150.0, "ly": 90.0, "cuts": sku_cuts.serialize(c),
        }],
    }
    wb = Workbook()
    wb.remove(wb.active)
    _build_by_sku(wb, contract, "Jul 2026")
    path = "/tmp/by_sku_unit.xlsx"
    wb.save(path)

    ws = load_workbook(path)["By SKU"]
    hdr = [ws.cell(4, i + 1).value for i in range(len(_SKU_HEADERS))]
    check("header count", len(hdr), len(_SKU_HEADERS))
    print(f"  {'PASS' if hdr == _SKU_HEADERS else 'FAIL'}  headers written verbatim")
    if hdr != _SKU_HEADERS:
        FAILURES.append("headers")

    # NB: several headers repeat by design ("% Share (of SKU)" appears in the
    # D2C, B2B, UK, US and ROW blocks), so a dict keyed by header collapses
    # them to the last occurrence. Read those by column index.
    row = {h: ws.cell(5, i + 1).value for i, h in enumerate(_SKU_HEADERS)}
    def col(header, block_start):
        """Index of `header` at or after 1-based column `block_start`."""
        return _SKU_HEADERS.index(header, block_start - 1) + 1
    check("net sales", row["Net Sales \u00a3"], 180.0)
    check("uk + us + row ties to total",
          row["UK TOTAL \u00a3"] + row["US TOTAL \u00a3"] + row["ROW TOTAL \u00a3"], 180.0)
    check("d2c uk", row["D2C UK \u00a3"], 100.0)
    check("us b2b units", row["B2B US Units"], 12)
    check("uk realised gm", row["UK Gross Margin %"], 1 - 20.0 / 100.0)
    check("us realised gm", row["US Gross Margin %"], 1 - 24.0 / 60.0)
    # ROW block's own "% Share (of SKU)" = 20 / 180
    check("row share of sku", ws.cell(5, col("% Share (of SKU)", 53)).value, 20.0 / 180.0)
    # D2C block's = 100+20 / 180 (D2C spans UK and ROW here)
    check("d2c share of sku", ws.cell(5, col("% Share (of SKU)", 21)).value, 120.0 / 180.0)
    check("material - finish", 1 if row["Material - Finish"] == "Brass - Antique Brass" else 0, 1)
    check("LM-1 falls back to lq when no prior contract", row["LM-1 Net Sales \u00a3"], 150.0)
    print(f"  wrote {path}")


def test_missing_cuts_falls_back_to_narrow_tab():
    """A pre-1.1 contract must render the narrow 12-column tab -- fully
    populated -- and say how to widen it. Not 58 empty columns, and not an
    exception either: rebuilding a published month's companion has to keep
    working, since those months can't be back-filled until their order
    exports are staged.
    """
    print("test_missing_cuts_falls_back_to_narrow_tab")
    from openpyxl import Workbook
    from excel_companion import _build_by_sku
    wb = Workbook(); wb.remove(wb.active)
    contract = {"contract_version": "1.0",
                "current": {"total_sales": 1.0, "units": 1, "d2c_gbp": 1.0, "b2b_gbp": 0.0,
                             "uk_gbp": 1.0, "us_gbp": 0.0, "row_gbp": 0.0},
                "skus_all": [{"sku": "OLD-001", "desc": "d", "coll": "c", "type_": "t",
                               "finish": "f", "gross": 1.0, "units": 1, "gm": 0.7,
                               "uk": 1.0, "us": 0.0, "uk_status": "Live"}]}
    ws = _build_by_sku(wb, contract, "Jun 2026")
    check("narrow tab width", ws.max_column, 12)
    note = str(ws.cell(ws.max_row, 1).value or "")
    ok = "backfill_sku_grain" in note
    print(f"  {'PASS' if ok else 'FAIL'}  note tells the reader how to widen it")
    if not ok:
        FAILURES.append("narrow-tab note")
    # every cell in the data row must be populated -- no blanks in the fallback
    blanks = [i for i in range(1, 13) if ws.cell(4, i).value is None]
    print(f"  {'PASS' if not blanks else 'FAIL'}  no blank cells in the narrow row "
          f"(blank cols: {blanks})")
    if blanks:
        FAILURES.append("narrow-tab blanks")


if __name__ == "__main__":
    for fn in (test_partition, test_realised_gm_differs_by_channel,
               test_unknown_cost_yields_none_not_zero, test_negative_and_zero_denominators,
               test_roundtrip, test_reconcile_gate_catches_a_leak,
               test_by_sku_tab_writes_and_ties, test_missing_cuts_falls_back_to_narrow_tab):
        fn()
        print()
    if FAILURES:
        print(f"FAILED: {len(FAILURES)} check(s): {FAILURES}")
        sys.exit(1)
    print("all checks passed")
