"""Monthly Trading builder -- revenue + country/channel + GM/inventory engine.

Sources order-lines directly from the Shopify Admin GraphQL API (see
shopify_feed.py) -- not ShopifyQL, not Matrixify, not Supermetrics. Real
orders give a stable line_item.id, so the order+SKU double-count trap
TRADING_logic_spec.md warns about (Q7, gap #4) doesn't apply here: every
line pulled is already atomic and uniquely identified.

STATUS (2026-08-03): structurally complete through the revenue + GM/inventory
aggregate. NOT yet passing the reconciliation gate -- reproduces May 2026
within ~1-5%, not the required 0.1% (see ROADMAP.md Phase B for the gap
writeup and leading suspects). The gate below correctly aborts and writes no
output until that closes; do not loosen the tolerance to force a pass.

Still to build once the gap closes: the full By SKU / By Collection
breakdown (needs Line Detail category/collection/finish enrichment), the
values-only workbook writer in the predecessor cell layout
(trading/dashboard/config.py), and the quarterly (Phase C) rollup.

Run:  python trading/build.py 2026-05
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook

from shopify_feed import STORES, fetch_store_orders
from revenue import country_bucket, channel, line_ab, refunds_by_line_id

from common.fx import ensure_dates, lookup as fx_lookup
from common.reconciliation_gate import assert_country_reconciles

FIXTURE_DIR = os.path.join(os.path.dirname(__file__), "tests", "fixtures")
TOL = 0.001


def fixture_totals(month_str):
    """Read UK/US/ROW/Total straight from a committed Monthly Trading Report
    fixture's Monthly Summary!AT7/CD7/DN7/F7 -- the regression oracle for
    this month, if one is committed. Returns None if no fixture exists for
    this month (nothing to regress against yet).
    """
    path = os.path.join(FIXTURE_DIR, f"{month_str}_Monthly_Trading_Report.xlsx")
    if not os.path.exists(path):
        return None
    wb = load_workbook(path, data_only=True)
    ws = wb["Monthly Summary"]
    return {
        "UK": ws["AT7"].value,
        "US": ws["CD7"].value,
        "ROW": ws["DN7"].value,
        "Total": ws["F7"].value,
    }


def _month_query(month_str):
    """'2026-05' -> Shopify search query for that calendar month."""
    year, month = (int(x) for x in month_str.split("-"))
    start = f"{year:04d}-{month:02d}-01"
    if month == 12:
        end = f"{year + 1:04d}-01-01"
    else:
        end = f"{year:04d}-{month + 1:02d}-01"
    return f"created_at:>={start} created_at:<{end}"


def pull_store_lines(store, month_str):
    """Pull every non-test order line for one store/month, with AB computed.

    Returns (lines, expected_order_count) where lines is a list of dicts:
    {sku, country, channel, ab, units, unit_cost}.
    """
    date_query = _month_query(month_str)
    orders, expected_count = fetch_store_orders(store, date_query)

    if store["label"] == "us":
        order_dates = [o["createdAt"][:10] for o in orders]
        fx_rows = ensure_dates(order_dates)

    lines = []
    for order in orders:
        ship_to = (order["shippingAddress"] or {}).get("countryCodeV2")
        bucket = country_bucket(ship_to, store["label"])
        chan = channel(order["purchasingEntity"])
        fx_rate = 1.0 if store["label"] == "uk" else fx_lookup(order["createdAt"][:10], fx_rows)
        refunds = refunds_by_line_id(order)

        for edge in order["lineItems"]["edges"]:
            li = edge["node"]
            net_of_discount = float(li["discountedTotalSet"]["shopMoney"]["amount"])
            tax = sum(float(t["priceSet"]["shopMoney"]["amount"]) for t in li["taxLines"])
            ref = refunds.get(li["id"], {"subtotal": 0.0, "tax": 0.0})
            returns_inc_vat = ref["subtotal"] + ref["tax"]
            tax_returned = ref["tax"]

            ab = line_ab(net_of_discount, tax, returns_inc_vat, tax_returned, fx_rate)

            variant = li.get("variant") or {}
            inv_item = (variant or {}).get("inventoryItem") or {}
            unit_cost_raw = (inv_item.get("unitCost") or {}).get("amount")

            lines.append({
                "sku": li["sku"],
                "country": bucket,
                "channel": chan,
                "ab": ab,
                "units": li["quantity"],
                "unit_cost": float(unit_cost_raw) if unit_cost_raw is not None else None,
            })

    return lines, expected_count, len(orders)


def aggregate_by_country(lines):
    totals = {"UK": 0.0, "US": 0.0, "ROW": 0.0}
    for line in lines:
        totals[line["country"]] += line["ab"]
    return totals


def build(month_str):
    all_lines = []
    row_count_report = {}
    for store in STORES:
        lines, expected_count, pulled_count = pull_store_lines(store, month_str)
        all_lines.extend(lines)
        row_count_report[store["label"]] = (pulled_count, expected_count)
        print(f"{store['label']}: pulled {pulled_count} orders (Shopify ordersCount: {expected_count})",
              file=sys.stderr)

    # Gate: row counts tie to Shopify's own totals (guards silent feed truncation)
    for label, (pulled, expected) in row_count_report.items():
        assert pulled == expected, (
            f"RECONCILE FAIL: {label} pulled {pulled} orders but Shopify ordersCount "
            f"reports {expected} -- pagination dropped rows"
        )

    country_totals = aggregate_by_country(all_lines)
    grand_total = sum(line["ab"] for line in all_lines)  # independent of country_totals
    assert_country_reconciles(country_totals, grand_total)

    block = {
        "UK": country_totals["UK"],
        "US": country_totals["US"],
        "ROW": country_totals["ROW"],
        "Total": grand_total,
    }

    fixture = fixture_totals(month_str)
    if fixture is not None:
        for label in ("UK", "US", "ROW", "Total"):
            computed, expected = block[label], fixture[label]
            rel = abs(computed - expected) / abs(expected) if expected else 0.0
            assert rel <= TOL, (
                f"RECONCILE FAIL: {label} computed {computed:,.2f} vs fixture "
                f"{expected:,.2f} (gap {rel:.4%}, tolerance {TOL:.1%}) -- see "
                f"ROADMAP.md Phase B known gap; run is NOT shipping a workbook"
            )
        print(f"reconciliation gate: PASS -- reproduces the {month_str} fixture within 0.1%",
              file=sys.stderr)
    else:
        print(f"reconciliation gate: no committed fixture for {month_str} -- "
              "internal checks only (row counts, uk+us+row)", file=sys.stderr)

    return block, all_lines


if __name__ == "__main__":
    month_str = sys.argv[1] if len(sys.argv) > 1 else "2026-05"
    block, lines = build(month_str)
    print(f"\n=== Monthly Trading revenue ({month_str}) ===")
    for label in ("UK", "US", "ROW", "Total"):
        print(f"  {label:6s} £{block[label]:,.2f}")
    print(f"\n{len(lines)} order lines processed.")
