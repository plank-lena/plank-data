"""Build the trading dashboard's Excel companion -- values-only, matching the
visual format of Plank_Q2_2026_Trade_Companion_REAL.xlsx (measured directly off
that file, see common/excel_styling.py). Called automatically from
build_matrixify_dashboard.py (monthly) and build_matrixify_quarterly_dashboard.py
(quarterly) -- not a separate manual step.

Generalized over period count so the same code serves both:
  - monthly:   constituent_contracts = [(period_label, current_contract)]  (len 1)
  - quarterly: constituent_contracts = [("Apr 2026", c1), ("May 2026", c2), ("Jun 2026", c3)]

current_contract is always the contract for the FULL reported period (the month's
own contract for monthly; the quarterly aggregate contract for quarterly) -- the
source for Summary/Drill-Down/By Collection/By SKU/Cuts. constituent_contracts
drives the By Period and Reconciliation tabs' per-period breakdown; with exactly
one entry, the redundant explicit total row is skipped (it would just repeat the
single period's own numbers).

Fixed 2026-08-13: `finishes` used to carry no per-finish GM or ROW (unlike
Product Type/Collection/SKU, which always did) -- GM was actually being
accumulated internally already (for the Finish Analysis GM% sidebar) but never
exposed; ROW wasn't tracked at all. Both now real, in contract.py's
emit_contract_from_matrixify -- see its finish_totals accumulation.
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from common.excel_styling import (
    title_row, subtitle_row, section_header, table_header, data_row, total_row,
    kpi_block, note_row, block_label, FMT_CURRENCY, FMT_PERCENT, FMT_PERCENT_SIGNED, FMT_INT,
)
from common import sku_cuts


def _build_summary(wb, cc, period_label, period_note):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16

    cur = cc["current"]
    total = cur["total_sales"]
    uk, us, row = cur["uk_gbp"], cur["us_gbp"], cur["row_gbp"]
    d2c, b2b = cur["d2c_gbp"], cur["b2b_gbp"]
    n_collections = sum(1 for c in cc["collections"] if c.get("ts", 0) > 0)

    title_row(ws, 1, f"{period_label} - Trading Report (Companion)", 8)
    subtitle_row(ws, 2, "REAL DATA - pipeline contract  |  ex-VAT  |  drill-down on 'Trade Drill-Down'", 8)
    note_row(ws, 3, f"Source: {period_label} committed trading contract(s). Values-only (no formulas).", 8)

    section_header(ws, 5, f"{period_label} Headline KPIs", 8)
    kpi_block(ws, 6, "Total Revenue", f"£{total:,.0f}", period_note, 1)
    kpi_block(ws, 6, "Units Sold", f"{cur['units']:,.0f}", f"{len(cc['skus_all'])} SKUs", 3)
    kpi_block(ws, 6, "Gross Margin", f"{cur['gm_pct']:.1%}", "revenue-weighted", 5)
    d2c_share = d2c / (d2c + b2b) if (d2c + b2b) else 0
    kpi_block(ws, 6, "D2C Share (of UK+US)", f"{d2c_share:.1%}", "channel excludes ROW", 7)

    kpi_block(ws, 10, "\U0001F1EC\U0001F1E7 UK Revenue", f"£{uk:,.0f}", f"{uk/total:.1%} of total" if total else "-", 1)
    kpi_block(ws, 10, "\U0001F1FA\U0001F1F8 US Revenue", f"£{us:,.0f}", f"{us/total:.1%} of total" if total else "-", 3)
    kpi_block(ws, 10, "\U0001F30D ROW Revenue", f"£{row:,.0f}", f"{row/total:.1%} of total" if total else "-", 5)
    kpi_block(ws, 10, "Collections", str(n_collections), "with sales this period", 7)

    section_header(ws, 15, "Revenue by Country  (reconciliation key)", 3)
    table_header(ws, 16, ["Country", "Revenue (ex-VAT)", "Share"])
    rows = [("United Kingdom", uk), ("United States", us), ("Rest of World (ROW)", row)]
    r = 17
    for name, val in rows:
        data_row(ws, r, [name, val, val / total if total else 0], [None, FMT_CURRENCY, FMT_PERCENT])
        r += 1
    total_row(ws, r, ["TOTAL", total, 1.0], [None, FMT_CURRENCY, FMT_PERCENT])
    r += 1
    note_row(ws, r, "Country is the reconciliation key. UK + US + ROW = Total. "
                     "ROW is never derived from the channel split.", 3)
    r += 2

    section_header(ws, r, "Revenue by Channel  (NOT a reconciliation key - excludes ROW)", 3)
    r += 1
    table_header(ws, r, ["Channel", "Revenue (ex-VAT)", "Share of UK+US"])
    r += 1
    uk_us = uk + us
    data_row(ws, r, ["D2C", d2c, d2c / uk_us if uk_us else 0], [None, FMT_CURRENCY, FMT_PERCENT])
    r += 1
    data_row(ws, r, ["B2B", b2b, b2b / uk_us if uk_us else 0], [None, FMT_CURRENCY, FMT_PERCENT])
    r += 1
    total_row(ws, r, ["UK + US subtotal", uk_us, 1.0], [None, FMT_CURRENCY, FMT_PERCENT])
    r += 1
    note_row(ws, r, f"Channel excludes ROW (£{row:,.0f}), which is carried inconsistently across "
                     f"channels. Never reconcile the headline from the channel split.", 3)


def _build_by_period(wb, constituent_contracts, cc, period_label):
    multi = len(constituent_contracts) > 1
    ws = wb.create_sheet("By Period" if not multi else "By Month")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJK", [16, 15, 12, 11, 13, 13, 12, 13, 12, 11, 11]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - " + ("Monthly Trajectory" if multi else "Period Detail"), 11)
    subtitle_row(ws, 2, "from the committed trading contract(s)  |  ex-VAT  |  vs LM / vs LY as reported", 11)

    headers = ["Period", "Revenue", "Units", "GM", "UK", "US", "ROW", "D2C", "B2B", "vs LM", "vs LY"]
    table_header(ws, 3, headers)
    fmts = [None, FMT_CURRENCY, FMT_INT, FMT_PERCENT, FMT_CURRENCY, FMT_CURRENCY,
            FMT_CURRENCY, FMT_CURRENCY, FMT_CURRENCY, FMT_PERCENT_SIGNED, FMT_PERCENT_SIGNED]

    r = 4
    for label, c in constituent_contracts:
        cur = c["current"]
        data_row(ws, r, [
            label, cur["total_sales"], cur["units"], cur["gm_pct"],
            cur["uk_gbp"], cur["us_gbp"], cur["row_gbp"], cur["d2c_gbp"], cur["b2b_gbp"],
            cur.get("vs_lm"), cur.get("vs_ly"),
        ], fmts)
        r += 1

    if multi:
        qcur = cc["current"]
        total_row(ws, r, [
            f"{period_label} TOTAL", qcur["total_sales"], qcur["units"], qcur["gm_pct"],
            qcur["uk_gbp"], qcur["us_gbp"], qcur["row_gbp"], qcur["d2c_gbp"], qcur["b2b_gbp"],
            None, None,
        ], fmts)
        r += 1
        note_row(ws, r, "vs LM / vs LY are month-on-month figures from each source contract; a "
                         "quarter-on-quarter comparative is not derivable at this grain, so it is omitted.", 11)


def _style_drill_level(ws, row, level, ncols=10):
    from common.excel_styling import DRILL_FILLS, DRILL_INDENT, DRILL_BOLD
    from openpyxl.styles import Alignment, PatternFill, Font
    fill = PatternFill("solid", fgColor=DRILL_FILLS[level])
    font_color = "FFFFFF" if level == 0 else "1F2933"
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Arial", size=10, bold=DRILL_BOLD[level], color=font_color)
        c.fill = fill
        if col == 1:
            c.alignment = Alignment(horizontal="left", vertical="center", indent=DRILL_INDENT[level])
        else:
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].outline_level = level
    if level == 3:
        ws.row_dimensions[row].hidden = True


def _build_drill_down(wb, cc, period_label):
    ws = wb.create_sheet("Trade Drill-Down")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False
    for col, w in zip("ABCDEFGHIJ", [34, 20, 15, 12, 10, 13, 13, 12, 13, 13]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - Trade Drill-Down", 10)
    subtitle_row(ws, 2, "Total -> Product Type -> Collection -> SKU  |  click the +/- outline "
                         "controls to expand  |  ex-VAT", 10)
    note_row(ws, 3, "SKUs collapsed by default - use the outline +/- to drill in. Children sum "
                     "to parents; UK+US+ROW = revenue at every level.", 10)

    headers = ["Product / Collection / SKU", "Type / Finish", "Revenue", "Units", "GM",
               "UK", "US", "ROW", "D2C", "B2B"]
    table_header(ws, 5, headers)
    fmts = [None, None, FMT_CURRENCY, FMT_INT, FMT_PERCENT, FMT_CURRENCY, FMT_CURRENCY,
            FMT_CURRENCY, FMT_CURRENCY, FMT_CURRENCY]

    cur = cc["current"]
    r = 6
    data_row(ws, r, ["TOTAL", "", cur["total_sales"], cur["units"], cur["gm_pct"],
                      cur["uk_gbp"], cur["us_gbp"], cur["row_gbp"], cur["d2c_gbp"], cur["b2b_gbp"]], fmts)
    _style_drill_level(ws, r, 0)
    r += 1

    skus_by_coll = {}
    for s in cc["skus_all"]:
        skus_by_coll.setdefault((s["type_"], s["coll"]), []).append(s)
    colls_by_type = {}
    for c in cc["collections"]:
        colls_by_type.setdefault(c["t"], []).append(c)

    for pt in sorted(cc["prod_types"], key=lambda x: -x["sales"]):
        data_row(ws, r, [pt["t"], f"{len(colls_by_type.get(pt['t'], []))} collections",
                          pt["sales"], pt["units"], pt["gm"], None, None, None, None, None], fmts)
        _style_drill_level(ws, r, 1)
        r += 1
        for coll in sorted(colls_by_type.get(pt["t"], []), key=lambda x: -x["ts"]):
            data_row(ws, r, [coll["c"], pt["t"], coll["ts"], coll["tu"], coll.get("gm"),
                              coll["uk_s"], coll["us_s"], coll["row_s"], coll["d2c"], coll["b2b"]], fmts)
            _style_drill_level(ws, r, 2)
            r += 1
            for s in sorted(skus_by_coll.get((pt["t"], coll["c"]), []), key=lambda x: -(x.get("gross") or 0)):
                data_row(ws, r, [s["sku"], s["finish"], s["gross"], s["units"], s["gm"],
                                  s["uk"], s["us"], (s["gross"] or 0) - (s["uk"] or 0) - (s["us"] or 0),
                                  s["d2c"], s["b2b"]], fmts)
                _style_drill_level(ws, r, 3)
                r += 1

    ws.freeze_panes = "A6"
    ws.sheet_format.outlineLevelRow = 3


def _build_by_collection(wb, cc, period_label):
    ws = wb.create_sheet("By Collection")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJ", [8, 20, 14, 14, 10, 12, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - By Collection (all)", 10)
    subtitle_row(ws, 2, "ranked by revenue  |  ex-VAT  |  filterable", 10)
    table_header(ws, 3, ["Rank", "Collection", "Type", "Revenue", "Units", "GM", "UK", "US", "ROW", "D2C"])
    fmts = [FMT_INT, None, None, FMT_CURRENCY, FMT_INT, FMT_PERCENT,
            FMT_CURRENCY, FMT_CURRENCY, FMT_CURRENCY, FMT_CURRENCY]

    r = 4
    for i, c in enumerate(sorted(cc["collections"], key=lambda c: -c["ts"]), start=1):
        data_row(ws, r, [i, c["c"], c["t"], c["ts"], c["tu"], c.get("gm"),
                          c["uk_s"], c["us_s"], c["row_s"], c["d2c"]], fmts)
        r += 1
    ws.freeze_panes = "A4"


# ── By SKU: the full column set ──────────────────────────────────────────
# Laid out to match the hand-built Monthly Trading Report's own By-SKU tab
# (2026-08-13, Lena) -- same block order, same block bands, same header
# wording, so someone reading both side by side finds each column where they
# expect it. The COLUMN LIST matches; the FIGURES deliberately do not, and
# that is the whole point: the report is gross-basis (returns not netted) and
# diverges from our per-SKU figures in both directions, so every number here
# is our own line_ab-derived value. Two intentional departures from the
# report's layout, both additive:
#
#   * A ROW block. The report has no ROW column, so its per-SKU UK + US
#     misses ROW revenue entirely (July 2026: GBP 9,359, 1.75% of the month).
#     Reproducing that would break the reconciliation contract at SKU grain.
#   * "Net Sales" rather than "Gross Sales" in the money headers. The report's
#     label is wrong -- the figure it names has returns netted out (locked
#     decision, trading_logic_spec.md) -- and carrying a wrong word into a new
#     deliverable propagates it. Same number, honest header.
#
# Columns the report carries that are NOT here, and why: "cc Size (mm)",
# "Available?", "Screw?", "IMG" and "US Supplier Cost incl Tariffs" are not
# parsed from the Line Detail master (see line_detail.COLUMN_MAP -- adding
# them is a one-line change per column IF the master carries those headers,
# which needs confirming against the live sheet, not assumed).

_SKU_BLOCKS = [
    ("", 1, 12),
    ("TOTAL", 13, 20),
    ("D2C", 21, 25),
    ("B2B", 26, 30),
    ("UK", 31, 41),
    ("US", 42, 52),
    ("ROW", 53, 56),
    ("LAST MONTH (LM-1)", 57, 63),
    ("LAST YEAR (LY LM)", 64, 70),
]

_SKU_HEADERS = [
    # attributes (1-12)
    "Ranking", "SKU", "Product Description", "Product Category", "Product Type",
    "Sub Category", "Material", "Finish", "Material - Finish", "Family/Collection",
    "UK Status", "US Status",
    # TOTAL (13-20)
    "Net Sales \u00a3", "TOTAL Units", "vs LM-1", "% Share of Total",
    "Inv Units", "Sell Through", "Weeks Cover", "Gross Margin %",
    # D2C (21-25)
    "D2C \u00a3", "D2C Units", "% Share (of SKU)", "% Share (of Channel)", "D2C Gross Margin %",
    # B2B (26-30)
    "B2B \u00a3", "B2B Units", "% Share (of SKU)", "% Share (of Channel)", "B2B Gross Margin %",
    # UK (31-41)
    "UK TOTAL \u00a3", "UK TOTAL Units", "% Share (of SKU)", "% Share (of Location)",
    "UK Gross Margin %", "D2C UK \u00a3", "D2C UK Units", "D2C UK Gross Margin %",
    "B2B UK \u00a3", "B2B UK Units", "B2B UK Gross Margin %",
    # US (42-52)
    "US TOTAL \u00a3", "US TOTAL Units", "% Share (of SKU)", "% Share (of Location)",
    "US Gross Margin %", "D2C US \u00a3", "D2C US Units", "D2C US Gross Margin %",
    "B2B US \u00a3", "B2B US Units", "B2B US Gross Margin %",
    # ROW (53-56)
    "ROW TOTAL \u00a3", "ROW TOTAL Units", "% Share (of SKU)", "% Share (of Location)",
    # LM-1 (57-63)
    "LM-1 Net Sales \u00a3", "LM-1 Units", "LM-1 D2C \u00a3", "LM-1 B2B \u00a3",
    "LM-1 UK \u00a3", "LM-1 US \u00a3", "LM-1 ROW \u00a3",
    # LY LM (64-70)
    "LY LM Net Sales \u00a3", "LY LM Units", "LY LM D2C \u00a3", "LY LM B2B \u00a3",
    "LY LM UK \u00a3", "LY LM US \u00a3", "LY LM ROW \u00a3",
]

_C, _I, _P, _S = FMT_CURRENCY, FMT_INT, FMT_PERCENT, FMT_PERCENT_SIGNED
_SKU_FORMATS = (
    [_I] + [None] * 11
    + [_C, _I, _S, _P, _I, _P, "0.0", _P]
    + [_C, _I, _P, _P, _P]
    + [_C, _I, _P, _P, _P]
    + [_C, _I, _P, _P, _P, _C, _I, _P, _C, _I, _P]
    + [_C, _I, _P, _P, _P, _C, _I, _P, _C, _I, _P]
    + [_C, _I, _P, _P]
    + [_C, _I, _C, _C, _C, _C, _C]
    + [_C, _I, _C, _C, _C, _C, _C]
)

_SKU_WIDTHS = {1: 8, 2: 22, 3: 34, 4: 16, 5: 14, 6: 16, 7: 14, 8: 16, 9: 22, 10: 16, 11: 12, 12: 12}


def _prior_cuts_index(prior_contract):
    """{sku: cuts} from a prior period's contract, or {} if that contract
    predates the SKU-grain fields. Used for the LM-1 and LY LM blocks: the
    prior period's own committed contract is the only source on the same
    basis as this period's figures, and chaining to it costs nothing (no
    re-derivation) while guaranteeing the comparison is against what was
    actually published.
    """
    if not prior_contract:
        return {}
    return {s["sku"]: sku_cuts.deserialize(s["cuts"])
            for s in prior_contract.get("skus_all", []) if s.get("cuts")}


def _build_by_sku_narrow(wb, cc, period_label):
    """The pre-2026-08-13 By-SKU tab: 12 columns, everything the contract holds
    at SKU grain before `cuts` existed. Retained deliberately, not as dead
    code -- it is what a version-1.0 contract (April/May/June 2026, and every
    2025 month) renders, so those months' companions stay reproducible without
    a back-fill. Back-fill a month and it gets the full tab instead.
    """
    ws = wb.create_sheet("By SKU")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJKL", [8, 20, 30, 14, 12, 13, 11, 12, 10, 11, 11, 11]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - By SKU (full population)", 12)
    subtitle_row(ws, 2, "ranked by revenue  |  ex-VAT  |  filterable", 12)
    table_header(ws, 3, ["Rank", "SKU", "Description", "Collection", "Type", "Finish",
                          "UK status", "Revenue", "Units", "GM", "UK", "US"])
    fmts = [FMT_INT, None, None, None, None, None, None,
            FMT_CURRENCY, FMT_INT, FMT_PERCENT, FMT_CURRENCY, FMT_CURRENCY]

    r = 4
    for i, s in enumerate(sorted(cc["skus_all"], key=lambda s: -(s.get("gross") or 0)), start=1):
        data_row(ws, r, [i, s["sku"], s["desc"], s["coll"], s["type_"], s["finish"],
                          s.get("uk_status"), s["gross"], s["units"], s["gm"], s["uk"], s["us"]], fmts)
        r += 1
    ws.freeze_panes = "A4"
    note_row(ws, r + 1, (
        f"This is the narrow By-SKU view: {period_label}'s contract predates the per-SKU "
        f"channel/country breakdown (contract_version "
        f"{cc.get('contract_version') or 'pre-1.1'!r}), so the D2C/B2B/UK/US/ROW cross, the "
        f"realised margins and the LM-1/LY LM blocks aren't available for it. To get the full "
        f"column set, back-fill the period: python trading/backfill_sku_grain.py "
        f"\"{period_label}\" --write. Nothing shown here is affected either way."), 12)
    return ws


def _build_by_sku(wb, cc, period_label, lm_contract=None, ly_contract=None):
    ws = wb.create_sheet("By SKU")
    ws.sheet_view.showGridLines = False
    ncols = len(_SKU_HEADERS)
    assert ncols == _SKU_BLOCKS[-1][2], "By SKU: header count and block map disagree"
    assert len(_SKU_FORMATS) == ncols, "By SKU: format count and header count disagree"

    for idx, w in _SKU_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = w
    for idx in range(13, ncols + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 15

    missing = [s["sku"] for s in cc["skus_all"] if not s.get("cuts")]
    if len(missing) == len(cc["skus_all"]):
        # A pre-1.1 contract holds none of the cuts. Write the narrow tab this
        # module always wrote rather than 58 empty columns OR an exception:
        # rebuilding a published month's companion must keep working, and a
        # short tab where every cell is populated beats a wide one where most
        # aren't. The note says how to widen it.
        wb.remove(ws)
        return _build_by_sku_narrow(wb, cc, period_label)

    title_row(ws, 1, f"{period_label} - By SKU", ncols)
    subtitle_row(ws, 2, "ranked by net sales  |  ex-VAT, net of in-window returns and "
                        "per-line discounts  |  filterable", ncols)
    for label, c0, c1 in _SKU_BLOCKS:
        if label:
            block_label(ws, 3, label, c0, c1)
    table_header(ws, 4, _SKU_HEADERS)

    lm_index = _prior_cuts_index(lm_contract)
    ly_index = _prior_cuts_index(ly_contract)

    cur = cc["current"]
    grand = {
        "total": cur["total_sales"], "d2c": cur["d2c_gbp"], "b2b": cur["b2b_gbp"],
        "uk": cur["uk_gbp"], "us": cur["us_gbp"], "row": cur["row_gbp"],
    }

    r = 5
    rows_written = 0
    for i, s in enumerate(sorted(cc["skus_all"], key=lambda s: -(s.get("gross") or 0)), start=1):
        cuts = sku_cuts.deserialize(s["cuts"]) if s.get("cuts") else None
        lm = lm_index.get(s["sku"])
        ly = ly_index.get(s["sku"])

        def cut(key, field="rev"):
            return cuts[key][field] if cuts else None

        def gm(key):
            return sku_cuts.gm_of(cuts[key]) if cuts else None

        def share_of_grand(key):
            den = grand.get(key)
            return (cut(key) / den) if (cuts and den) else None

        def share_of_sku(key):
            tot = s.get("gross")
            return (cut(key) / tot) if (cuts and tot and tot > 0) else None

        material = s.get("material")
        finish = s.get("finish")
        mat_finish = f"{material} - {finish}" if (material and finish) else (material or finish or None)

        values = [
            i, s["sku"], s["desc"], s.get("item_type"), s["type_"], s.get("style"),
            material, finish, mat_finish, s["coll"], s.get("uk_status"), s.get("us_status"),
            # TOTAL
            s["gross"], s["units"], s.get("vslq"),
            (s["gross"] / grand["total"]) if grand["total"] else None,
            s.get("inv"), s.get("st"), s.get("wc"), gm("total") if cuts else s.get("gm"),
            # D2C / B2B
            cut("d2c"), cut("d2c", "u"), share_of_sku("d2c"), share_of_grand("d2c"), gm("d2c"),
            cut("b2b"), cut("b2b", "u"), share_of_sku("b2b"), share_of_grand("b2b"), gm("b2b"),
            # UK
            cut("uk"), cut("uk", "u"), share_of_sku("uk"), share_of_grand("uk"), gm("uk"),
            cut("uk_d2c"), cut("uk_d2c", "u"), gm("uk_d2c"),
            cut("uk_b2b"), cut("uk_b2b", "u"), gm("uk_b2b"),
            # US
            cut("us"), cut("us", "u"), share_of_sku("us"), share_of_grand("us"), gm("us"),
            cut("us_d2c"), cut("us_d2c", "u"), gm("us_d2c"),
            cut("us_b2b"), cut("us_b2b", "u"), gm("us_b2b"),
            # ROW
            cut("row"), cut("row", "u"), share_of_sku("row"), share_of_grand("row"),
            # LM-1 -- from the prior period's own contract, never re-derived here
            lm["total"]["rev"] if lm else s.get("lq"),
            lm["total"]["u"] if lm else None,
            lm["d2c"]["rev"] if lm else None, lm["b2b"]["rev"] if lm else None,
            lm["uk"]["rev"] if lm else None, lm["us"]["rev"] if lm else None,
            lm["row"]["rev"] if lm else None,
            # LY LM
            ly["total"]["rev"] if ly else s.get("ly"),
            ly["total"]["u"] if ly else None,
            ly["d2c"]["rev"] if ly else None, ly["b2b"]["rev"] if ly else None,
            ly["uk"]["rev"] if ly else None, ly["us"]["rev"] if ly else None,
            ly["row"]["rev"] if ly else None,
        ]
        data_row(ws, r, values, _SKU_FORMATS)
        r += 1
        rows_written += 1

    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{r - 1}"

    note = (f"{rows_written} SKUs with sales in {period_label}. Figures are this pipeline's own, "
            f"net of in-window returns and per-line discounts, ex-VAT -- they will NOT match the "
            f"hand-built Monthly Trading Report cell for cell, which reports on a gross basis. "
            f"LM-1 and LY LM blocks come from those periods' committed contracts, so they are on "
            f"the same basis as the current period.")
    if missing:
        note += (f" {len(missing)} SKU(s) have no channel/country breakdown: their published "
                 f"figures no longer tie to a re-derivation, so the cuts were left out rather "
                 f"than estimated (see provenance.sku_cuts_backfill.unmatched).")
    note_row(ws, r + 1, note, ncols)


def _build_cuts(wb, cc, period_label):
    ws = wb.create_sheet("Cuts")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [24, 16, 14, 12, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - Dimensional Cuts", 7)
    subtitle_row(ws, 2, "product status / product type / finish  |  ex-VAT", 7)
    cur = cc["current"]
    fmts = [None, FMT_CURRENCY, FMT_INT, FMT_PERCENT, FMT_CURRENCY, FMT_CURRENCY, FMT_CURRENCY]

    def cut_table(start_row, title, col_label, rows, key_name, sales_key, units_key, gm_key, uk_key, us_key, row_key):
        r = start_row
        section_header(ws, r, title, 7)
        r += 1
        table_header(ws, r, [col_label, "Revenue", "Units", "GM", "UK", "US", "ROW"])
        r += 1
        for item in rows:
            data_row(ws, r, [item[key_name], item[sales_key], item[units_key],
                              item.get(gm_key) if gm_key else None,
                              item.get(uk_key) if uk_key else None,
                              item.get(us_key) if us_key else None,
                              item.get(row_key) if row_key else None], fmts)
            r += 1
        total_row(ws, r, ["TOTAL", cur["total_sales"], cur["units"], cur["gm_pct"],
                           cur["uk_gbp"], cur["us_gbp"], cur["row_gbp"]], fmts)
        return r + 2

    r = 4
    r = cut_table(r, "By Product Status (SKU UK status)", "Status",
                  sorted(cc["statuses"], key=lambda x: -x["sales"]),
                  "s", "sales", "units", "gm", None, None, None)
    r = cut_table(r, "By Product Type", "Product type",
                  sorted(cc["prod_types"], key=lambda x: -x["sales"]),
                  "t", "sales", "units", "gm", None, None, None)
    finishes_rows = [{"name": k, **v} for k, v in
                      sorted(cc["finishes"].items(), key=lambda kv: -kv[1]["total"])]
    r = cut_table(r, "By Finish", "Finish", finishes_rows,
                  "name", "total", "units", "gm", "uk", "us", "row")


def _build_reconciliation(wb, constituent_contracts, cc, period_label):
    multi = len(constituent_contracts) > 1
    ws = wb.create_sheet("Reconciliation")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [26, 18, 18, 14, 14]):
        ws.column_dimensions[col].width = w

    title_row(ws, 1, f"{period_label} - Trade Reconciliation", 5)
    subtitle_row(ws, 2, "uk + us + row -> total, per period  |  ex-VAT", 5)

    section_header(ws, 4, "Country reconciliation by period", 5)
    table_header(ws, 5, ["Period", "UK", "US", "ROW", "Diff vs total"])
    fmts = [None, FMT_CURRENCY, FMT_CURRENCY, FMT_CURRENCY, FMT_CURRENCY]
    r = 6
    for label, c in constituent_contracts:
        cur = c["current"]
        diff = (cur["uk_gbp"] + cur["us_gbp"] + cur["row_gbp"]) - cur["total_sales"]
        data_row(ws, r, [label, cur["uk_gbp"], cur["us_gbp"], cur["row_gbp"], diff], fmts)
        r += 1
    if multi:
        qcur = cc["current"]
        qdiff = (qcur["uk_gbp"] + qcur["us_gbp"] + qcur["row_gbp"]) - qcur["total_sales"]
        data_row(ws, r, [period_label, qcur["uk_gbp"], qcur["us_gbp"], qcur["row_gbp"], qdiff], fmts)
        r += 1
    r += 1

    qcur = cc["current"]
    qsum = qcur["uk_gbp"] + qcur["us_gbp"] + qcur["row_gbp"]
    rel_diff = abs(qsum - qcur["total_sales"]) / qcur["total_sales"] if qcur["total_sales"] else 0
    passed = rel_diff <= 0.001 and qcur["row_gbp"] != 0
    data_row(ws, r, [f"{period_label} sum (uk+us+row)", qsum], [None, FMT_CURRENCY]); r += 1
    data_row(ws, r, [f"{period_label} headline total", qcur["total_sales"]], [None, FMT_CURRENCY]); r += 1
    data_row(ws, r, ["Relative difference", rel_diff], [None, FMT_PERCENT]); r += 1
    data_row(ws, r, ["Tolerance", 0.001], [None, FMT_PERCENT]); r += 1
    data_row(ws, r, ["ROW bucket present", f"YES (£{qcur['row_gbp']:,.0f})" if qcur["row_gbp"] else "NO"]); r += 1
    note_row(ws, r, ("PASS" if passed else "FAIL") +
             " - uk+us+row reconciles to total within 0.1%; live ROW bucket", 5)
    r += 2

    section_header(ws, r, "Notes", 5); r += 1
    notes = [
        "Source: the committed trading contract(s) (trading/contracts/*.json) -- the pipeline's "
        "own reconciliation-gated output, not re-derived here.",
        f"{period_label} = " + ("sum of its constituent months. " if multi else "a single period. ") +
        "SKU-level ROW is carried explicitly (2026-08-13), not inferred as gross - UK - US; "
        "uk + us + row ties to each SKU's own total as well as to the headline.",
        "Country is the reconciliation key, never channel. All ex-VAT; returns reported separately.",
        "Values-only (no formulas).",
    ]
    for n in notes:
        note_row(ws, r, f"\u2022  {n}", 5)
        r += 1


def build_companion(out_path, period_label, current_contract, constituent_contracts,
                    period_note="", lm_contract=None, ly_contract=None):
    """The one function callers need. See module docstring for the two calling
    shapes (monthly: 1 constituent; quarterly: 3).

    lm_contract / ly_contract: the prior-month and prior-year contracts as
    already-loaded dicts, optional. They feed the By-SKU tab's LM-1 and LY LM
    blocks at SKU grain. Passing them is strictly additive -- omit them and
    those blocks fall back to the contract's own per-SKU lq/ly totals, which
    is what every caller did before 2026-08-13.
    """
    wb = Workbook()
    _build_summary(wb, current_contract, period_label, period_note)
    _build_by_period(wb, constituent_contracts, current_contract, period_label)
    _build_drill_down(wb, current_contract, period_label)
    _build_by_collection(wb, current_contract, period_label)
    _build_by_sku(wb, current_contract, period_label, lm_contract=lm_contract,
                  ly_contract=ly_contract)
    _build_cuts(wb, current_contract, period_label)
    _build_reconciliation(wb, constituent_contracts, current_contract, period_label)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
