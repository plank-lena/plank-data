"""Extract all data from the monthly trading xlsx into structured dicts."""

import re
from pathlib import Path

import openpyxl

from config import (
    MS_ROW7, LM_BLOCK, LY_BLOCK, PERIOD_CELLS,
    STATUS_COLS, TYPE_COLS,
    FINISH_COLS, COLL_COL, COLL_COL_Q, SKU_COL,
)


# ── Period model ──────────────────────────────────────────────────────────────

def _parse_period(value):
    """Parse 'May - 2026' → {label:'May 2026', short:"May '26"}."""
    m = re.match(r'(\w+)\s*-\s*(\d{4})', str(value).strip())
    if not m:
        raise ValueError(f'Cannot parse period cell: {value!r}')
    month, year = m.group(1), m.group(2)
    return {'label': f'{month} {year}', 'short': f"{month[:3]} '{year[2:]}"}


def extract_period_model(ws):
    """Return period model dict with keys cm/lm/ly each having label+short."""
    pm = {}
    for key, cell in PERIOD_CELLS.items():
        pm[key] = _parse_period(ws[cell].value)
    return pm


# ── Monthly Summary headline (row 7) ─────────────────────────────────────────

def _cell(ws, col_letter, row=7):
    return ws[f'{col_letter}{row}'].value


def extract_headline(ws):
    """Return dict of all row-7 values from Monthly Summary."""
    h = {k: _cell(ws, col) for k, col in MS_ROW7.items()}
    lm = {k: _cell(ws, col) for k, col in LM_BLOCK.items()}
    ly = {k: _cell(ws, col) for k, col in LY_BLOCK.items()}
    return {'current': h, 'lm': lm, 'ly': ly}


# ── Product Status rows ───────────────────────────────────────────────────────

def extract_statuses(ws):
    """Dynamically discover every Product Status row (BRIEF step 5: the
    same fixed-row-dict bug class as Product Type/Finish -- the previous
    hardcoded STATUS_ROWS only read Continuity/Newness/Discontinued/Dead
    (rows 8-11) and silently dropped 'Not For Sale' and 'Pre-Launch' (rows
    12-13), which carry real revenue in both the monthly and quarterly
    sheets (e.g. £2,453.99 'Not For Sale' in the May fixture -- previously
    documented as "no row in the statuses table at all", which was wrong;
    it has a row, the row was just never read). Found while building the
    quarterly aggregator, whose correctness depends on summing complete
    monthly status data.
    """
    header_row = None
    for row in range(1, ws.max_row + 1):
        if ws[f'B{row}'].value == 'Product Status':
            header_row = row
            break
    if header_row is None:
        raise ValueError("extract_statuses: 'Product Status' header row not found")

    rows = []
    r = header_row + 1
    name = ws[f'B{r}'].value
    if str(name).strip().upper() == 'TOTAL':
        r += 1  # skip the TOTAL/headline row itself -- statuses are the breakdown below it
    while True:
        name = ws[f'B{r}'].value
        if name is None:
            break
        row = {'s': str(name).strip()}
        for field, col in STATUS_COLS.items():
            row[field] = ws[f'{col}{r}'].value
        rows.append(row)
        r += 1
    return rows


# ── Product Type rows ─────────────────────────────────────────────────────────

def extract_product_types(ws):
    """Dynamically discover every Product Type block from the Monthly
    Summary 'Product Type | Product Category' table -- a department's row
    is marked by Product Category == 'TOTAL'; every row after it (until the
    next TOTAL row or the table's blank terminator) is one of its
    subcategories (BRIEF #4 step 4 §5/§10: retire the fixed row dict --
    it silently hid Taps and Door, which only ever had £0 in the frozen May
    snapshot but must render like any other department once real data
    exists for them).
    """
    header_row = None
    for row in range(1, ws.max_row + 1):
        if ws[f'B{row}'].value == 'Product Type' and ws[f'C{row}'].value == 'Product Category':
            header_row = row
            break
    if header_row is None:
        raise ValueError("extract_product_types: 'Product Type' header row not found")

    depts = []
    current = None
    r = header_row + 1
    while True:
        b = ws[f'B{r}'].value
        c = ws[f'C{r}'].value
        if b is None and c is None:
            break
        if c == 'TOTAL':
            if current is not None:
                depts.append(current)
            current = {
                't':       str(b).strip(),
                'sales':   ws[f'F{r}'].value or 0,
                'units':   ws[f'J{r}'].value or 0,
                'vs_lq':   ws[f'G{r}'].value,
                'vs_ly':   ws[f'I{r}'].value,
                'gm':      ws[f'S{r}'].value,
                'subcats': [],
            }
        elif current is not None:
            current['subcats'].append({
                'name':  str(c).strip(),
                'sales': ws[f'F{r}'].value or 0,
                'units': ws[f'J{r}'].value or 0,
                'vs_ly': ws[f'I{r}'].value,
            })
        r += 1
    if current is not None:
        depts.append(current)
    return depts


# ── Finish rows ───────────────────────────────────────────────────────────────

def extract_finishes(ws):
    """Dynamically discover every named Finish row (BRIEF #4 step 4 §5/§10:
    the previous fixed 8-row curation only ever read a hand-picked subset --
    the real sheet carries ~29 named finishes. Whatever finishes exist in
    the data render; a consumer that wants a curated top-N does that as a
    display cut, not by dropping rows here.
    """
    header_row = None
    for row in range(1, ws.max_row + 1):
        if ws[f'B{row}'].value == 'Finish':
            header_row = row
            break
    if header_row is None:
        raise ValueError("extract_finishes: 'Finish' header row not found")

    finishes = {}
    r = header_row + 1
    while True:
        name = ws[f'B{r}'].value
        if name is None:
            break
        f = {}
        for field, col in FINISH_COLS.items():
            f[field] = ws[f'{col}{r}'].value
        finishes[str(name).strip()] = f
        r += 1
    return finishes


# ── By Collection ─────────────────────────────────────────────────────────────

def _detect_coll_layout(ws_coll):
    """Detect which By-Collection column layout this SHEET actually has,
    rather than trusting mode='month'/'quarter'. Found while building the
    Step 5 quarterly aggregator: the April 2026 monthly export is missing
    the 'vs LY LM' column entirely -- the exact same one-column shift from
    that point on already documented for the quarterly layout (COLL_COL_Q)
    -- even though April is a genuine MONTHLY report. Column layout is a
    property of which sheet-template vintage produced this specific export,
    not of month-vs-quarter; trusting mode alone silently corrupted every
    UK/US/ROW figure for April (reading UK Units into the UK £ slot, etc.).
    """
    header_row = None
    for row in range(1, min(ws_coll.max_row, 10) + 1):
        vals = [c.value for c in ws_coll[row]]
        if any(v == 'UK £' for v in vals if isinstance(v, str)):
            header_row = row
            break
    if header_row is None:
        raise ValueError("extract_collections: could not find the 'UK £' header row to detect column layout")
    header_vals = [c.value for c in ws_coll[header_row]]
    return COLL_COL if 'vs LY LM' in header_vals else COLL_COL_Q


def extract_collections(ws_coll, ws_sku, mode='month'):
    """Return list of collection dicts (all rows with gross > 0).

    The column mapping is DETECTED from the sheet's own header row (see
    _detect_coll_layout), not chosen from `mode` -- some monthly exports
    use the same shifted layout the quarterly sheet uses. `mode` is kept
    as a parameter for backward compatibility but no longer drives this
    choice.
    """
    # Build SKU count per (type, coll) from By SKU
    sku_counts = {}
    for row in ws_sku.iter_rows(min_row=5, values_only=True):
        if row[SKU_COL['sku']] is None:
            continue
        key = (str(row[SKU_COL['type_']] or '').strip(),
               str(row[SKU_COL['coll']] or '').strip())
        sku_counts[key] = sku_counts.get(key, 0) + 1

    collections = []
    C = _detect_coll_layout(ws_coll)
    for row in ws_coll.iter_rows(min_row=5, values_only=True):
        if row[C['rank']] is None:
            continue
        gross = row[C['gross']]
        if not gross or gross <= 0:
            continue
        # A handful of real rows (e.g. ALERIA, a blank-typed BECKER distinct
        # from Cabinetry's BECKER) never got a Product Type populated in the
        # sheet -- same "Unknown" convention as an unmatched SKU's
        # department elsewhere, not a blank/empty department key.
        t = str(row[C['type_']] or '').strip() or 'Unknown'
        c = str(row[C['coll']] or '').strip()
        uk_s = row[C['uk']] or 0
        us_s = row[C['us']] or 0
        lq_total = row[C['lq_total']] or 0
        lq_uk = row[C['lq_uk']] or 0
        lq_us = row[C['lq_us']] or 0
        uk_vs = (uk_s - lq_uk) / lq_uk if lq_uk else None
        us_vs = (us_s - lq_us) / lq_us if lq_us else None
        collections.append({
            'r':        int(row[C['rank']]),
            't':        t,
            'c':        c,
            'ts':       gross,
            'tu':       int(row[C['units']] or 0),
            'vs_lq':    row[C['vs_lq']],
            'gm':       row[C['gm']],
            'st':       row[C['st']],
            'wc':       row[C['wc']],
            'd2c':      row[C['d2c']] or 0,
            'b2b':      row[C['b2b']] or 0,
            'uk_s':     uk_s,
            'us_s':     us_s,
            'row_s':    row[C['row']] or 0,
            'lq_total': lq_total,
            'lq_uk':    lq_uk,
            'lq_us':    lq_us,
            'uk_vs':    uk_vs,
            'us_vs':    us_vs,
            'skus':     sku_counts.get((t, c), 0),
        })
    return collections


# ── By SKU ────────────────────────────────────────────────────────────────────

def extract_skus_all(ws_sku):
    """Return list of ALL SKU dicts from By SKU sheet (row 5+)."""
    skus = []
    C = SKU_COL
    for row in ws_sku.iter_rows(min_row=5, values_only=True):
        if row[C['sku']] is None:
            continue
        gross = row[C['gross']]
        if not isinstance(gross, (int, float)) or gross <= 0:
            continue
        skus.append({
            'rank':      row[C['rank']],
            'sku':       str(row[C['sku']]).strip(),
            'desc':      str(row[C['desc']] or '').strip(),
            'coll':      str(row[C['coll']] or '').strip(),
            # Same 'Unknown' default as extract_collections' blank-type_
            # rows (ALERIA / a blank-typed BECKER / BOBBIN / CANTO) -- both
            # sheets must agree on a SKU's department, or the completeness
            # tripwire (validate._completeness_errors) flags a false gap.
            'type_':     str(row[C['type_']] or '').strip() or 'Unknown',
            'finish':    str(row[C['finish']] or '').strip(),
            'uk_status': str(row[C['uk_status']] or '').strip(),
            'us_status': str(row[C['us_status']] or '').strip(),
            'gross':     gross,
            'units':     int(row[C['units']] or 0),
            'vslq':      row[C['vslq']],
            'gm':        row[C['gm']],
            'st':        row[C['st']],
            'wc':        row[C['wc']],
            'inv':       int(row[C['inv']] or 0) if row[C['inv']] is not None else 0,
            'd2c':       row[C['d2c']] or 0,
            'b2b':       row[C['b2b']] or 0,
            'uk':        row[C['uk']] or 0,
            'uk_u':      int(row[C['uk_u']] or 0),
            'us':        row[C['us']] or 0,
            'us_u':      int(row[C['us_u']] or 0),
            'lq':        row[C['lq']],
            'ly':        row[C['ly']],
        })
    return skus


# ── Master entry point ────────────────────────────────────────────────────────

def extract_all(path):
    """Load xlsx and return complete raw data dict.

    Auto-detects monthly vs quarterly source by summary sheet name
    ('Monthly Summary' vs 'Quarterly Summary'). The two layouts are
    otherwise structurally identical except for the By-Collection column
    shift handled in extract_collections/COLL_COL_Q.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    if 'Quarterly Summary' in wb.sheetnames:
        ws_ms = wb['Quarterly Summary']
        mode = 'quarter'
    else:
        ws_ms = wb['Monthly Summary']
        mode = 'month'
    ws_coll = wb['By Collection']
    ws_sku  = wb['By SKU']

    headline = extract_headline(ws_ms)
    return {
        'mode':         mode,
        'period_model': extract_period_model(ws_ms),
        'current':      headline['current'],
        'lm':           headline['lm'],
        'ly':           headline['ly'],
        'statuses':     extract_statuses(ws_ms),
        'prod_types':   extract_product_types(ws_ms),
        'finishes':     extract_finishes(ws_ms),
        'collections':  extract_collections(ws_coll, ws_sku, mode=mode),
        'skus_all':     extract_skus_all(ws_sku),
        '_ws_ms':       ws_ms,
        '_ws_coll':     ws_coll,
        '_ws_sku':      ws_sku,
    }
