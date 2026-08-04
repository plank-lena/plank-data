"""Monthly Trading builder -- Matrixify frozen-snapshot source.

Supersedes build.py's live Shopify GraphQL source (shopify_feed.py) for the
reasons in ROADMAP.md Phase B / the 2026-08-03 Matrixify migration brief:
live queries gave timezone-boundary leakage, mid-run drift against an
actively-changing store, and non-reproducibility. This reuses the SAME
AB/country/channel/reconciliation-gate logic (revenue.py, common/) against
rows parsed from a committed Matrixify export CSV instead.

STATUS (2026-08-03): both Matrixify-PlankUK and Matrixify-PlankUS are
connected. UK+US now reconcile via compute_combined() below (BRIEF #5:
ROW bucket + three-way country reconcile) -- ship-to country, NOT store of
origin, is the bucketing key, so a UK-store order shipped to e.g. Ireland
correctly lands in ROW rather than being counted as UK. compute_store()
below (single-store, store-label-only reporting) is kept for the per-store
diagnostic CLI use it already had; it is not the reconciled figure -- its
own "Total" is a store's grand total across all three of ITS buckets, not
that store's UK/US/ROW component alone. Use compute_combined() (or the
`reconcile` CLI action) for anything compared against the oracle.

The US £ residual (~-6.5%, standalone Discount rows / cancelled-order
handling) is explicitly OUT of scope here -- see trading/RECONCILE_HANDOFF.md.
Do not tune country bucketing to chase it.

Run:  python trading/build_matrixify.py trading/source/orders_2026-05_US.csv us 2026-05
      python trading/build_matrixify.py reconcile 2026-05
"""
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from matrixify_source import load_rows, build_lines, order_month_london
from revenue import country_bucket, line_ab

from common.fx import ensure_month, seed_confirmed, lookup_month
from common.reconciliation_gate import assert_country_reconciles, assert_matches_oracle

JULY_FX_RATE = 1.3250
JULY_FX_SOURCE = "googlefinance (sheet AA, month rate)"

# May 2026 Monthly Summary row 7 ground truth (AT7/CD7/DN7/F7/AX7/CH7/DR7/J7),
# restated here rather than imported from build.py's identical MAY_TOTALS --
# build.py is the superseded live-query path scheduled for retirement (see
# ROADMAP.md Phase B / RECONCILE_HANDOFF.md), so this module should not take
# a dependency on it just to share a constant.
MAY_THREE_WAY = {
    "total": 476292.70, "uk": 247772.02, "us": 214063.73, "row": 14456.95,
    "units_total": 27453, "units_uk": 16373, "units_us": 9436, "units_row": 822,
}


def channel_from_company(company):
    """B2B if a Company is present, else D2C -- see brief Step 3. Channel
    never partitions the reconciled total (ROADMAP.md §4).
    """
    return "B2B" if company else "D2C"


def _fx_rate_for(store_label, month_str):
    if month_str == "2026-07":
        fx_rows = seed_confirmed(month_str, JULY_FX_RATE, JULY_FX_SOURCE)
    else:
        fx_rows = ensure_month(month_str)
    return 1.0 if store_label == "uk" else lookup_month(month_str, fx_rows)


def _load_store_month(csv_path, store_label, month_str):
    """Load one store's export and filter to the lines belonging to
    month_str. Returns (kept_lines, fx_rate, shipping_total, order_names).
    Shared by compute_store() (single-store diagnostic) and
    compute_combined() (the reconciled ship-to-country three-way) so the
    loading/filtering logic can't drift between the two.
    """
    rows = load_rows(csv_path)
    lines, orders_meta, order_shipping = build_lines(rows, store_label)
    fx_rate = _fx_rate_for(store_label, month_str)

    kept_lines = [line for line in lines if line["order_month"] == month_str]
    dropped_other_month = len(lines) - len(kept_lines)
    order_names = {line["order_name"] for line in kept_lines}

    shipping_total = sum(
        v for name, v in order_shipping.items()
        if orders_meta.get(name, {}).get("created_at")
        and order_month_london(orders_meta[name]["created_at"]) == month_str
    ) / fx_rate

    print(f"{store_label}: {len(rows)} raw CSV rows, {len(kept_lines)} lines kept for "
          f"{month_str} ({dropped_other_month} lines belonged to a different month), "
          f"{len(order_names)} distinct orders", file=sys.stderr)

    return kept_lines, fx_rate, shipping_total, order_names


def compute_store(csv_path, store_label, month_str):
    """Single-store diagnostic: buckets one store's own lines by ship-to
    country. NOT the reconciled figure -- its "Total" is that store's grand
    total across all three of ITS OWN buckets (e.g. the UK store's own
    ROW-shipped orders are included), not a UK/US/ROW component to compare
    against the oracle on its own. Use compute_combined() for that.
    """
    kept_lines, fx_rate, shipping_total, order_names = _load_store_month(csv_path, store_label, month_str)

    country_totals = {"UK": 0.0, "US": 0.0, "ROW": 0.0}
    units_totals = {"UK": 0, "US": 0, "ROW": 0}
    channel_totals = {"D2C": 0.0, "B2B": 0.0}
    tax_total = 0.0
    returns_total = 0.0

    for line in kept_lines:
        bucket = country_bucket(line["ship_country_code"], store_label)
        chan = channel_from_company(line["company"])
        ab = line_ab(line["net_of_discount"], line["tax"], line["returns_inc_vat"],
                     line["tax_returned"], fx_rate)

        country_totals[bucket] += ab
        units_totals[bucket] += line["units"]
        channel_totals[chan] += ab
        tax_total += line["tax"] / fx_rate
        returns_total += line["returns_inc_vat"] / fx_rate

    return {
        "country_totals": country_totals,
        "units_totals": units_totals,
        "channel_totals": channel_totals,
        "tax_total": tax_total,
        "returns_total": returns_total,
        "shipping_total": shipping_total,
        "grand_total": sum(country_totals.values()),
        "order_count": len(order_names),
    }


def compute_combined(uk_csv, us_csv, month_str, include_returns=True):
    """Union UK + US store lines for month_str and bucket EVERY line by
    SHIP-TO country (revenue.country_bucket: GB->UK, US->US, else->ROW,
    falling back to the line's own store only when ship-to is blank/N-A --
    never to ROW). This is the reconciled three-way figure BRIEF #5 wants:
    a UK-store order shipped to Ireland lands in ROW, not UK.

    grand_total is accumulated independently of country_totals (a separate
    running sum over every kept line, not sum(country_totals.values())) so
    assert_country_reconciles's uk+us+row check is a real leak check, not
    vacuous -- see its docstring in common/reconciliation_gate.py.

    include_returns: DIAGNOSTIC ONLY (default True, matches today's shipped
    behaviour). False computes the GROSS line value (net_of_discount - tax)
    / fx_rate, i.e. the returns term dropped entirely, still ex-VAT and
    FX-converted -- see floor_isolation_test_matrixify() and
    RECONCILE_HANDOFF.md's returns-netting basis question. Units are
    unaffected either way: line["units"] already comes from the original
    Line Item row's own Quantity, never netted against Refund Line rows.
    """
    country_totals = {"UK": 0.0, "US": 0.0, "ROW": 0.0}
    units_totals = {"UK": 0, "US": 0, "ROW": 0}
    channel_totals = {"D2C": 0.0, "B2B": 0.0}
    tax_total = 0.0
    returns_total = 0.0
    shipping_total = 0.0
    grand_total = 0.0
    order_names = set()

    for csv_path, store_label in ((uk_csv, "uk"), (us_csv, "us")):
        kept_lines, fx_rate, store_shipping, store_order_names = _load_store_month(
            csv_path, store_label, month_str)
        shipping_total += store_shipping
        # prefix so an order name can never collide across the two stores
        order_names |= {f"{store_label}:{name}" for name in store_order_names}

        for line in kept_lines:
            bucket = country_bucket(line["ship_country_code"], store_label)
            chan = channel_from_company(line["company"])
            if include_returns:
                ab = line_ab(line["net_of_discount"], line["tax"], line["returns_inc_vat"],
                             line["tax_returned"], fx_rate)
            else:
                ab = (line["net_of_discount"] - line["tax"]) / fx_rate

            country_totals[bucket] += ab
            units_totals[bucket] += line["units"]
            channel_totals[chan] += ab
            tax_total += line["tax"] / fx_rate
            returns_total += line["returns_inc_vat"] / fx_rate
            grand_total += ab

    return {
        "country_totals": country_totals,
        "units_totals": units_totals,
        "channel_totals": channel_totals,
        "tax_total": tax_total,
        "returns_total": returns_total,
        "shipping_total": shipping_total,
        "grand_total": grand_total,
        "order_count": len(order_names),
    }


def report_combined(result, oracle, label_prefix):
    ct, ut = result["country_totals"], result["units_totals"]
    print(f"\n=== {label_prefix} three-way reconcile (ship-to country) ===")
    print(f"{'bucket':10s} {'computed':>15s} {'oracle':>15s} {'gap %':>8s}")
    for key, label in (("uk", "UK"), ("us", "US"), ("row", "ROW")):
        computed = ct[label]
        expected = oracle[key]
        gap = abs(computed - expected) / abs(expected) * 100 if expected else float("nan")
        print(f"{label:10s} {computed:15,.2f} {expected:15,.2f} {gap:7.3f}%")
    total_gap = abs(result["grand_total"] - oracle["total"]) / abs(oracle["total"]) * 100
    print(f"{'Total':10s} {result['grand_total']:15,.2f} {oracle['total']:15,.2f} {total_gap:7.3f}%")

    print("\n--- units (diagnostic only -- NOT gated; oracle's own UK+US+ROW units "
          "don't foot to its Total units column, a sheet-side quirk, not our bug) ---")
    for key, label in (("units_uk", "UK"), ("units_us", "US"), ("units_row", "ROW")):
        print(f"  {label:6s} {ut[label]:>8,.0f}  (oracle {oracle[key]:,})")
    units_total = sum(ut.values())
    print(f"  {'Total':6s} {units_total:>8,.0f}  (oracle {oracle['units_total']:,})")


def floor_isolation_test_matrixify(uk_csv, us_csv, month_str, oracle):
    """Diagnostic (2026-08-04): decisive test for the returns-netting basis
    question raised in RECONCILE_HANDOFF.md -- is the UK -5.16% / US -6.47%
    May shortfall a returns-netting basis error (oracle is gross, builder
    computes net-of-returns), or genuinely order-scope?

    Computes GROSS (include_returns=False: ex-VAT, FX-converted, returns
    term dropped) and NET (include_returns=True: today's shipped AB output)
    against the same oracle row-7 targets. Units are reported too, but they
    are already gross either way (see compute_combined's docstring) -- a
    remaining units gap under the GROSS branch points to a residual
    order-scope component, not a returns-basis one.

    Never called from the gate path -- report-only, no assertion, so it
    can run against any month without needing a committed oracle fixture
    for it (the caller supplies oracle).
    """
    gross = compute_combined(uk_csv, us_csv, month_str, include_returns=False)
    net = compute_combined(uk_csv, us_csv, month_str, include_returns=True)

    print(f"\n=== Floor isolation ({month_str}): GROSS (returns term dropped) vs NET (shipped) vs oracle ===")
    print(f"{'bucket':8s} {'GROSS':>15s} {'NET':>15s} {'oracle':>15s} {'GROSS gap':>10s} {'NET gap':>10s}")
    for key, label in (("uk", "UK"), ("us", "US"), ("row", "ROW")):
        g, n, o = gross["country_totals"][label], net["country_totals"][label], oracle[key]
        g_gap = (g - o) / o * 100 if o else float("nan")
        n_gap = (n - o) / o * 100 if o else float("nan")
        print(f"{label:8s} {g:15,.2f} {n:15,.2f} {o:15,.2f} {g_gap:+9.3f}% {n_gap:+9.3f}%")
    g_tot, n_tot, o_tot = gross["grand_total"], net["grand_total"], oracle["total"]
    g_gap = (g_tot - o_tot) / o_tot * 100
    n_gap = (n_tot - o_tot) / o_tot * 100
    print(f"{'Total':8s} {g_tot:15,.2f} {n_tot:15,.2f} {o_tot:15,.2f} {g_gap:+9.3f}% {n_gap:+9.3f}%")

    print("\n--- units (identical under GROSS/NET -- never netted against Refund Line rows) ---")
    gu = gross["units_totals"]
    for key, label in (("units_uk", "UK"), ("units_us", "US"), ("units_row", "ROW")):
        u_gap = (gu[label] - oracle[key]) / oracle[key] * 100 if oracle.get(key) else float("nan")
        print(f"  {label:6s} {gu[label]:>8,.0f}  (oracle {oracle[key]:,}, gap {u_gap:+.3f}%)")
    units_total = sum(gu.values())
    ut_gap = (units_total - oracle["units_total"]) / oracle["units_total"] * 100
    print(f"  {'Total':6s} {units_total:>8,.0f}  (oracle {oracle['units_total']:,}, gap {ut_gap:+.3f}%)")

    return gross, net


def _read_oracle_row7(xlsx_path):
    """Read one monthly oracle's 'Monthly Summary' row 7 (the TOTAL-status
    row) ground truth, generalising MAY_THREE_WAY's hardcoded columns
    (F/AT/AX/DN=6/46/82/118 revenue, J/AZ/CH/DR-equivalent=10/50/86/122
    units) to any committed monthly oracle fixture. data_only=True to read
    resolved values, not formulas.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Monthly Summary"]
    row = 7
    assert ws.cell(row=row, column=2).value == "TOTAL", (
        f"{xlsx_path}: row 7 col B is not 'TOTAL' -- oracle layout has shifted, do not trust these numbers")
    return {
        "total": ws.cell(row=row, column=6).value, "uk": ws.cell(row=row, column=46).value,
        "us": ws.cell(row=row, column=82).value, "row": ws.cell(row=row, column=118).value,
        "units_total": ws.cell(row=row, column=10).value, "units_uk": ws.cell(row=row, column=50).value,
        "units_us": ws.cell(row=row, column=86).value, "units_row": ws.cell(row=row, column=122).value,
    }


def recomposition_diagnostic(month_str, uk_csv, us_csv, oracle):
    """Report-only (2026-08-04): the recomposition diagnostic from
    RECONCILE_HANDOFF.md's floor-isolation follow-up. For one month, per
    ship-to bucket, computes three revenue variants and vs the oracle:

      A = GROSS: (Line: Total - Line: Tax Total) / fx_rate summed over
          Line Item rows -- per-line discounts applied (Line: Total already
          nets those), no order-level discount codes, no returns removed.
      B = A - order-level `Discount` line-type rows (own Line: Total,
          already negative, /fx_rate), bucketed by the owning order's
          ship-to country.
      C = B - the same two terms for orders carrying a non-blank
          `Cancelled At` (a blanket exclusion, per RECONCILE_HANDOFF.md's
          "29-order scope" question -- NOT a refined Payment: Status rule).

    Also reports, independently from scratch (not reusing build_lines()'s
    output), four unit counts per bucket: the builder's own unit count,
    Sigma Line: Quantity over Line Item rows only (should match the
    builder exactly -- a from-scratch cross-check, not a second opinion),
    Sigma |Refund Line quantity| (returned units, currently untracked
    anywhere in the shipped pipeline), and the oracle's Gross Unit Sold.

    Never called from the gate path. Returns the raw numbers so
    run_recomposition_diagnostic() can check the cross-month anti-
    overfitting bar (B~oracle for UK, C~oracle for US, across all three
    months) without re-parsing the CSVs.
    """
    buckets = ("UK", "US", "ROW")
    A = dict.fromkeys(buckets, 0.0)
    order_discount = dict.fromkeys(buckets, 0.0)
    cancelled_a = dict.fromkeys(buckets, 0.0)
    cancelled_discount = dict.fromkeys(buckets, 0.0)
    cancelled_orders = dict.fromkeys(buckets, 0)
    line_item_units = dict.fromkeys(buckets, 0.0)
    refund_units = dict.fromkeys(buckets, 0.0)

    # "builder unit count (as shipped)" -- the ACTUAL production compute_combined()
    # path, not a duplicate of this diagnostic's own formula, so the units table
    # below is a real cross-check of the shipped pipeline, not a tautology.
    builder_units = dict(compute_combined(uk_csv, us_csv, month_str)["units_totals"])

    for csv_path, store_label in ((uk_csv, "uk"), (us_csv, "us")):
        rows = load_rows(csv_path)
        fx_rate = _fx_rate_for(store_label, month_str)

        orders_meta = {}
        for row in rows:
            if row.get("Top Row", "").lower() == "true":
                orders_meta[row["Name"]] = {
                    "created_at": row["Created At"],
                    "cancelled_at": row.get("Cancelled At") or None,
                    "ship_country_code": row.get("Shipping: Country Code") or None,
                }

        def kept(name, _meta=orders_meta):
            meta = _meta.get(name)
            return meta is not None and meta["created_at"] and order_month_london(meta["created_at"]) == month_str

        order_a = defaultdict(float)
        order_disc = defaultdict(float)
        order_builder_units = defaultdict(float)
        order_refund_units = defaultdict(float)
        by_line = defaultdict(list)

        for row in rows:
            name = row["Name"]
            if not kept(name):
                continue
            line_type = row.get("Line: Type")
            if line_type == "Discount":
                order_disc[name] += float(row.get("Line: Total") or 0)
            elif line_type in ("Line Item", "Refund Line"):
                line_id = row.get("Line: ID")
                if line_id:
                    by_line[(name, line_id)].append(row)

        for (name, _line_id), line_rows in by_line.items():
            original = next((r for r in line_rows if r["Line: Type"] == "Line Item"), None)
            if original is None:
                continue
            net = float(original.get("Line: Total") or 0)
            tax = float(original.get("Line: Tax Total") or 0)
            qty = float(original.get("Line: Quantity") or 0)
            order_a[name] += (net - tax) / fx_rate
            order_builder_units[name] += qty
            order_refund_units[name] += sum(
                abs(float(r.get("Line: Quantity") or 0)) for r in line_rows if r["Line: Type"] == "Refund Line")

        for name, meta in orders_meta.items():
            if not kept(name):
                continue
            bucket = country_bucket(meta["ship_country_code"], store_label)
            a_val = order_a.get(name, 0.0)
            disc_val = order_disc.get(name, 0.0) / fx_rate

            A[bucket] += a_val
            order_discount[bucket] += disc_val
            line_item_units[bucket] += order_builder_units.get(name, 0.0)
            refund_units[bucket] += order_refund_units.get(name, 0.0)

            if meta["cancelled_at"] is not None:
                cancelled_a[bucket] += a_val
                cancelled_discount[bucket] += disc_val
                cancelled_orders[bucket] += 1

    B = {b: A[b] + order_discount[b] for b in buckets}
    C = {b: B[b] - cancelled_a[b] - cancelled_discount[b] for b in buckets}
    A["Total"] = sum(A[b] for b in buckets)
    B["Total"] = sum(B[b] for b in buckets)
    C["Total"] = sum(C[b] for b in buckets)

    def _gap(computed, expected):
        return (computed - expected) / expected * 100 if expected else float("nan")

    print(f"\n=== Recomposition diagnostic ({month_str}): A/B/C vs oracle ===")
    print(f"{'bucket':8s} {'A=GROSS':>14s} {'B=A-disc':>14s} {'C=B-cancl':>14s} {'oracle':>14s} "
          f"{'A gap':>9s} {'B gap':>9s} {'C gap':>9s}  cancelled orders")
    for key, label in (("uk", "UK"), ("us", "US"), ("row", "ROW")):
        o = oracle[key]
        print(f"{label:8s} {A[label]:14,.2f} {B[label]:14,.2f} {C[label]:14,.2f} {o:14,.2f} "
              f"{_gap(A[label], o):+8.3f}% {_gap(B[label], o):+8.3f}% {_gap(C[label], o):+8.3f}%  {cancelled_orders[label]}")
    o_tot = oracle["total"]
    print(f"{'Total':8s} {A['Total']:14,.2f} {B['Total']:14,.2f} {C['Total']:14,.2f} {o_tot:14,.2f} "
          f"{_gap(A['Total'], o_tot):+8.3f}% {_gap(B['Total'], o_tot):+8.3f}% {_gap(C['Total'], o_tot):+8.3f}%  "
          f"{sum(cancelled_orders.values())}")

    print("\n--- units: builder (as shipped) vs from-scratch Line Item recount vs returned vs oracle ---")
    for key, label in (("units_uk", "UK"), ("units_us", "US"), ("units_row", "ROW")):
        o = oracle[key]
        print(f"  {label:6s} builder {builder_units[label]:>8,.0f}  Line-Item-recount {line_item_units[label]:>8,.0f}  "
              f"returned {refund_units[label]:>8,.0f}  oracle {o:>8,.0f}  "
              f"(oracle-builder gap {_gap(o, builder_units[label]) if builder_units[label] else float('nan'):+.3f}% "
              f"of builder; oracle-builder vs returned: {o - builder_units[label]:+,.0f} vs {refund_units[label]:,.0f})")
    bt, lt, rt = sum(builder_units.values()), sum(line_item_units.values()), sum(refund_units.values())
    ot = oracle["units_total"]
    print(f"  {'Total':6s} builder {bt:>8,.0f}  Line-Item-recount {lt:>8,.0f}  returned {rt:>8,.0f}  oracle {ot:>8,.0f}  "
          f"(oracle-builder gap {(ot - bt):+,.0f} vs returned {rt:,.0f})")

    return {
        "A": A, "B": B, "C": C, "cancelled_orders": cancelled_orders,
        "builder_units": builder_units, "line_item_units": line_item_units, "refund_units": refund_units,
    }


def run_recomposition_diagnostic():
    """Runs recomposition_diagnostic() for April/May/June 2026 against
    their committed oracle fixtures and Matrixify exports, then checks the
    cross-month anti-overfitting bar from RECONCILE_HANDOFF.md: B~oracle
    for UK across all three months, C~oracle for US across all three
    months, and C must not push UK negative. Three components (line-
    discounts, order-discounts, cancelled-orders) is enough freedom to fit
    one month by luck -- this is the check that it isn't luck.
    """
    months = [
        ("2026-04", "trading/source/orders_2026-04_UK.csv", "trading/source/orders_2026-04_US.csv",
         "trading/tests/fixtures/2026-04_Monthly_Trading_Report.xlsx"),
        ("2026-05", "trading/source/orders_2026-05_UK.csv", "trading/source/orders_2026-05_US.csv",
         "trading/tests/fixtures/2026-05_Monthly_Trading_Report.xlsx"),
        ("2026-06", "trading/source/orders_2026-06_UK.csv", "trading/source/orders_2026-06_US.csv",
         "trading/tests/fixtures/2026-06_Monthly_Trading_Report.xlsx"),
    ]

    results = {}
    for month_str, uk_csv, us_csv, oracle_xlsx in months:
        oracle = _read_oracle_row7(oracle_xlsx)
        results[month_str] = recomposition_diagnostic(month_str, uk_csv, us_csv, oracle)

    def _gap(computed, expected):
        return (computed - expected) / expected * 100 if expected else float("nan")

    print("\n=== Cross-month anti-overfitting check ===")
    print(f"{'month':10s} {'B UK gap':>10s} {'C US gap':>10s} {'C UK gap':>10s}")
    for month_str, _, _, oracle_xlsx in months:
        oracle = _read_oracle_row7(oracle_xlsx)
        r = results[month_str]
        b_uk_gap = _gap(r["B"]["UK"], oracle["uk"])
        c_us_gap = _gap(r["C"]["US"], oracle["us"])
        c_uk_gap = _gap(r["C"]["UK"], oracle["uk"])
        print(f"{month_str:10s} {b_uk_gap:+9.3f}% {c_us_gap:+9.3f}% {c_uk_gap:+9.3f}%")

    return results


def _month_end_london(month_str):
    """'2026-06' -> the last calendar date of that month (date object)."""
    year, month = (int(x) for x in month_str.split("-"))
    next_month = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
    return (next_month - timedelta(days=1)).date()


def _parse_refund_dt(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d %H:%M:%S %z")


def maturity_cutoff_diagnostic(month_str, uk_csv, us_csv, oracle, cutoff_days_list):
    """Report-only (2026-08-05): tests the report-generation-timing/maturity
    hypothesis directly against our own committed Matrixify exports -- no
    live-sheet access needed. Matrixify's `Refund: Created At` column gives
    each refund event's own processing date, which the shipped pipeline
    never reads (matrixify_source.py's build_lines() sums all refund rows
    present in the export with no cutoff). Lena confirmed (2026-08-05) the
    live sheet's own April/May Returns breakdown is no longer obtainable
    (read-only, months no longer loaded) and that reports are generated
    ~9-15 days after month-end -- this reproduces that same maturity-window
    idea from data we already have, for all three committed months.

    For each candidate cutoff (days after the order-month's own calendar
    month-end), recomputes GROSS-minus-cutoff-returns (deliberately no
    order-discount term, to test returns-maturity in isolation) per bucket
    and reports the gap to oracle.
    """
    month_end = _month_end_london(month_str)
    buckets = ("UK", "US", "ROW")
    oracle_key = {"UK": "uk", "US": "us", "ROW": "row", "Total": "total"}

    gross = dict.fromkeys(buckets, 0.0)
    refund_events = {b: [] for b in buckets}  # [(days_after_month_end, gbp_amount), ...]

    for csv_path, store_label in ((uk_csv, "uk"), (us_csv, "us")):
        rows = load_rows(csv_path)
        fx_rate = _fx_rate_for(store_label, month_str)

        orders_meta = {}
        for row in rows:
            if row.get("Top Row", "").lower() == "true":
                orders_meta[row["Name"]] = {
                    "created_at": row["Created At"],
                    "ship_country_code": row.get("Shipping: Country Code") or None,
                }

        def kept(name, _meta=orders_meta):
            meta = _meta.get(name)
            return meta is not None and meta["created_at"] and order_month_london(meta["created_at"]) == month_str

        by_line = defaultdict(list)
        for row in rows:
            name = row["Name"]
            if not kept(name):
                continue
            if row.get("Line: Type") in ("Line Item", "Refund Line"):
                line_id = row.get("Line: ID")
                if line_id:
                    by_line[(name, line_id)].append(row)

        for (name, _line_id), line_rows in by_line.items():
            original = next((r for r in line_rows if r["Line: Type"] == "Line Item"), None)
            if original is None:
                continue
            bucket = country_bucket(orders_meta[name]["ship_country_code"], store_label)
            net = float(original.get("Line: Total") or 0)
            tax = float(original.get("Line: Tax Total") or 0)
            gross[bucket] += (net - tax) / fx_rate

            for r in line_rows:
                if r["Line: Type"] != "Refund Line":
                    continue
                refund_dt = _parse_refund_dt(r.get("Refund: Created At"))
                if refund_dt is None:
                    continue
                days_after = (refund_dt.date() - month_end).days
                amount_gbp = float(r.get("Line: Total") or 0) / fx_rate  # already negative
                refund_events[bucket].append((days_after, amount_gbp))

    gross["Total"] = sum(gross[b] for b in buckets)

    print(f"\n=== Maturity-cutoff diagnostic ({month_str}): GROSS minus cutoff-returns, "
          f"NO discount term ===")
    header = f"{'cutoff':>7s} " + " ".join(f"{b + ' gap':>10s}" for b in buckets) + f" {'Total gap':>10s}"
    print(header)
    for cutoff in cutoff_days_list:
        gaps = []
        for b in buckets + ("Total",):
            if b == "Total":
                returns_in_window = sum(amt for bb in buckets for days, amt in refund_events[bb] if days <= cutoff)
            else:
                returns_in_window = sum(amt for days, amt in refund_events[b] if days <= cutoff)
            computed = gross[b] + returns_in_window
            o = oracle[oracle_key[b]]
            gaps.append((computed - o) / o * 100 if o else float("nan"))
        print(f"{cutoff:>6d}d " + " ".join(f"{g:+9.3f}%" for g in gaps[:-1]) + f" {gaps[-1]:+9.3f}%")

    n_events = {b: len(refund_events[b]) for b in buckets}
    print(f"  refund events seen (any cutoff): {n_events}, total {sum(n_events.values())}")
    return {"gross": gross, "refund_events": refund_events}


def gate_check_combined(result, oracle):
    """ROW present + uk+us+row == independent grand total, then each bucket
    (and the total) vs the committed May oracle within 0.1%. Raises on
    failure -- never called with output already written; the caller must
    gate before writing anything downstream.
    """
    assert_country_reconciles(result["country_totals"], result["grand_total"])
    assert_matches_oracle(
        {
            "uk": result["country_totals"]["UK"],
            "us": result["country_totals"]["US"],
            "row": result["country_totals"]["ROW"],
            "total": result["grand_total"],
        },
        oracle,
    )
    print("\ngate: ROW bucket present; uk+us+row == independent grand total; "
          "each bucket + total within 0.1% of the May oracle. PASS", file=sys.stderr)


if __name__ == "__main__":
    if sys.argv[1] == "maturity_cutoff":
        # python trading/build_matrixify.py maturity_cutoff
        months = [
            ("2026-04", "trading/source/orders_2026-04_UK.csv", "trading/source/orders_2026-04_US.csv",
             "trading/tests/fixtures/2026-04_Monthly_Trading_Report.xlsx"),
            ("2026-05", "trading/source/orders_2026-05_UK.csv", "trading/source/orders_2026-05_US.csv",
             "trading/tests/fixtures/2026-05_Monthly_Trading_Report.xlsx"),
            ("2026-06", "trading/source/orders_2026-06_UK.csv", "trading/source/orders_2026-06_US.csv",
             "trading/tests/fixtures/2026-06_Monthly_Trading_Report.xlsx"),
        ]
        cutoffs = [0, 1, 3, 5, 7, 9, 10, 11, 12, 13, 15, 20, 30, 45, 60, 90]
        for month_str, uk_csv, us_csv, oracle_xlsx in months:
            oracle = _read_oracle_row7(oracle_xlsx)
            maturity_cutoff_diagnostic(month_str, uk_csv, us_csv, oracle, cutoffs)
    elif sys.argv[1] == "recomposition":
        # python trading/build_matrixify.py recomposition
        run_recomposition_diagnostic()
    elif sys.argv[1] == "floor_isolation":
        # python trading/build_matrixify.py floor_isolation [month] [uk_csv] [us_csv]
        month_str = sys.argv[2] if len(sys.argv) > 2 else "2026-05"
        uk_csv = sys.argv[3] if len(sys.argv) > 3 else "trading/source/orders_2026-05_UK.csv"
        us_csv = sys.argv[4] if len(sys.argv) > 4 else "trading/source/orders_2026-05_US.csv"
        oracle = MAY_THREE_WAY if month_str == "2026-05" else None
        if oracle is None:
            print(f"floor_isolation: no committed oracle wired up for {month_str} yet", file=sys.stderr)
            sys.exit(1)
        floor_isolation_test_matrixify(uk_csv, us_csv, month_str, oracle)
    elif sys.argv[1] == "reconcile":
        # python trading/build_matrixify.py reconcile [month] [uk_csv] [us_csv]
        month_str = sys.argv[2] if len(sys.argv) > 2 else "2026-05"
        uk_csv = sys.argv[3] if len(sys.argv) > 3 else "trading/source/orders_2026-05_UK.csv"
        us_csv = sys.argv[4] if len(sys.argv) > 4 else "trading/source/orders_2026-05_US.csv"

        result = compute_combined(uk_csv, us_csv, month_str)
        if month_str == "2026-05":
            report_combined(result, MAY_THREE_WAY, "May 2026")
            gate_check_combined(result, MAY_THREE_WAY)  # raises + aborts before any output is written on failure
        else:
            print(f"\n=== {month_str} three-way (Matrixify source, ship-to country) ===")
            for label in ("UK", "US", "ROW"):
                print(f"  {label:6s} £{result['country_totals'][label]:,.2f}")
            print(f"  Total  £{result['grand_total']:,.2f}")
            print(f"  (no committed oracle for {month_str} yet -- gate not run)")
    else:
        csv_path = sys.argv[1]
        store_label = sys.argv[2]
        month_str = sys.argv[3]

        result = compute_store(csv_path, store_label, month_str)
        print(f"\n=== {store_label} {month_str} (Matrixify source) ===")
        for label in ("UK", "US", "ROW"):
            print(f"  {label:6s} £{result['country_totals'][label]:,.2f}")
        print(f"  Total  £{result['grand_total']:,.2f}")
        print(f"  Units  {sum(result['units_totals'].values()):,}")
        print(f"  Orders {result['order_count']:,}")
        print(f"  Tax    £{result['tax_total']:,.2f}")
        print(f"  Returns £{-result['returns_total']:,.2f}")
        print(f"  Shipping £{result['shipping_total']:,.2f}")
